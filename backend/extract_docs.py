import pdfplumber
import openpyxl

# --- Extract PDF ---
pdf_lines = []
pdf = pdfplumber.open('Sample_table_of_contents.pdf')
for i, page in enumerate(pdf.pages):
    text = page.extract_text()
    if text:
        pdf_lines.append(f"--- PAGE {i+1} ---")
        for line in text.split('\n'):
            stripped = line.strip()
            if stripped:
                pdf_lines.append(stripped)
pdf.close()

# --- Extract Excel ---
excel_lines = []
wb = openpyxl.load_workbook("Project Report Submission DOs & Dont_s.xlsx")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    excel_lines.append(f"--- SHEET: {sheet_name} ---")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = [str(c) if c is not None else '' for c in row]
        line = ' | '.join(vals)
        if line.replace('|','').strip():
            excel_lines.append(line)
wb.close()

# Write output as UTF-8
with open('extracted_docs_utf8.txt', 'w', encoding='utf-8') as f:
    f.write("PDF CONTENT\n")
    f.write("="*60 + "\n")
    f.write('\n'.join(pdf_lines))
    f.write("\n\nEXCEL CONTENT\n")
    f.write("="*60 + "\n")
    f.write('\n'.join(excel_lines))

print("Done. Written to extracted_docs_utf8.txt")
