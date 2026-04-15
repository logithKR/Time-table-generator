from fastapi import FastAPI, Depends, HTTPException, BackgroundTasks
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
import json
import math
import io
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
import models, schemas
from database import engine, get_db
from fastapi.middleware.cors import CORSMiddleware


# Ensure tables exist
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="BIT Scheduler Pro (Simplified)")

# CORS
origins = [
    "http://localhost:5173", "http://127.0.0.1:5173",
    "http://localhost:5174", "http://127.0.0.1:5174",
    "http://localhost:5175", "http://127.0.0.1:5175",
    "https://timetable.bitsathy.ac.in",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/import-book1")
def run_import_book1(background_tasks: BackgroundTasks):
    background_tasks.add_task(import_book1.import_book1_data)
    return {"message": "Data import from Book1.xlsx has started in the background."}

@app.post("/api/sync-cms")
def run_sync_cms():
    from sync_cms_to_local import sync_databases
    try:
        sync_databases()
        return {"status": "success", "message": "CMS data synchronized successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")


# ============================================
# DEPARTMENTS
# ============================================

# ============================================
# SEMESTER CONFIG
# ============================================

@app.get("/api/semester-config", response_model=List[schemas.SemesterConfigResponse])
def get_semester_configs(db: Session = Depends(get_db)):
    return db.query(models.SemesterConfig).all()

@app.post("/api/semester-config/{semester}", response_model=schemas.SemesterConfigResponse)
def update_semester_config(semester: int, req: schemas.SemesterConfigUpdate, db: Session = Depends(get_db)):
    config = db.query(models.SemesterConfig).filter_by(semester=semester).first()
    if not config:
        config = models.SemesterConfig(semester=semester, academic_year=req.academic_year)
        db.add(config)
    else:
        config.academic_year = req.academic_year
    db.commit()
    db.refresh(config)
    return config

@app.get("/departments", response_model=List[schemas.Department])
def get_departments(db: Session = Depends(get_db)):
    return db.query(models.DepartmentMaster).all()

@app.post("/departments")
def create_department(req: schemas.DepartmentCreate, db: Session = Depends(get_db)):
    existing = db.query(models.DepartmentMaster).filter_by(department_code=req.department_code).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Department {req.department_code} already exists")
    dept = models.DepartmentMaster(
        department_code=req.department_code,
        student_count=req.student_count,
        pair_add_course_miniproject=req.pair_add_course_miniproject
    )
    db.add(dept)
    db.commit()
    return {"status": "success", "department_code": req.department_code}

@app.put("/departments/{code}", response_model=schemas.Department)
def update_department(code: str, req: schemas.DepartmentUpdate, db: Session = Depends(get_db)):
    dept = db.query(models.DepartmentMaster).filter_by(department_code=code).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    
    if req.student_count is not None:
        dept.student_count = req.student_count
    
    if req.pair_add_course_miniproject is not None:
        dept.pair_add_course_miniproject = req.pair_add_course_miniproject
        
    db.commit()
    db.refresh(dept)
    return dept

@app.get("/departments/{code}/capacities", response_model=List[schemas.DepartmentSemesterCountResponse])
def get_department_capacities(code: str, db: Session = Depends(get_db)):
    dept = db.query(models.DepartmentMaster).filter_by(department_code=code).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    
    return db.query(models.DepartmentSemesterCount).filter_by(department_code=code).order_by(models.DepartmentSemesterCount.semester).all()

@app.post("/departments/{code}/capacities", response_model=schemas.DepartmentSemesterCountResponse)
def upsert_department_capacity(
    code: str, 
    semester: int, 
    req: schemas.DepartmentSemesterCountUpdate, 
    db: Session = Depends(get_db)
):
    dept = db.query(models.DepartmentMaster).filter_by(department_code=code).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    
    record = db.query(models.DepartmentSemesterCount).filter_by(
        department_code=code, semester=semester
    ).first()
    
    if record:
        record.student_count = req.student_count
    else:
        record = models.DepartmentSemesterCount(
            department_code=code,
            semester=semester,
            student_count=req.student_count
        )
        db.add(record)
        
    db.commit()
    db.refresh(record)
    return record

@app.delete("/departments/{code}")
def delete_department(code: str, db: Session = Depends(get_db)):
    dept = db.query(models.DepartmentMaster).filter_by(department_code=code).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    # Check for dependent records
    fac_count = db.query(models.FacultyMaster).filter_by(department_code=code).count()
    course_count = db.query(models.CourseMaster).filter_by(department_code=code).count()
    if fac_count > 0 or course_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {fac_count} faculty and {course_count} courses still linked.")
    db.delete(dept)
    db.commit()
    return {"status": "deleted"}

@app.get("/semesters")
def get_semesters():
    return [{"semester_number": i} for i in range(1, 9)]

# ============================================
# FACULTY
# ============================================

@app.get("/faculty", response_model=List[schemas.Faculty])
def get_faculty(department_code: str = None, db: Session = Depends(get_db)):
    query = db.query(models.FacultyMaster)
    if department_code:
        query = query.filter(models.FacultyMaster.department_code == department_code)
    return query.all()

@app.post("/faculty")
def create_faculty(req: schemas.FacultyCreate, db: Session = Depends(get_db)):
    existing = db.query(models.FacultyMaster).filter_by(faculty_id=req.faculty_id).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Faculty {req.faculty_id} already exists")
    # Ensure department exists
    dept = db.query(models.DepartmentMaster).filter_by(department_code=req.department_code).first()
    if not dept:
        raise HTTPException(status_code=400, detail=f"Department {req.department_code} does not exist")
    fac = models.FacultyMaster(
        faculty_id=req.faculty_id,
        faculty_name=req.faculty_name,
        faculty_email=req.faculty_email,
        department_code=req.department_code,
        status=req.status
    )
    db.add(fac)
    db.commit()
    return {"status": "success", "faculty_id": req.faculty_id}

@app.delete("/faculty/{fid}")
def delete_faculty(fid: str, db: Session = Depends(get_db)):
    fac = db.query(models.FacultyMaster).filter_by(faculty_id=fid).first()
    if not fac:
        raise HTTPException(status_code=404, detail="Faculty not found")
    # Remove course-faculty mappings first
    db.query(models.CourseFacultyMap).filter_by(faculty_id=fid).delete()
    db.delete(fac)
    db.commit()
    return {"status": "deleted"}

# ============================================
# COURSES
# ============================================

@app.get("/courses", response_model=List[schemas.Course])
def get_courses(department_code: str = None, semester: int = None, db: Session = Depends(get_db)):
    query = db.query(models.CourseMaster)
    if department_code:
        query = query.filter(models.CourseMaster.department_code == department_code)
    if semester:
        query = query.filter(models.CourseMaster.semester == semester)
    return query.all()

@app.post("/courses")
def create_course(req: schemas.CourseCreate, db: Session = Depends(get_db)):
    target_depts = [req.department_code]
    if getattr(req, 'common_departments', None):
        for d in req.common_departments:
            if d not in target_depts:
                target_depts.append(d)

    for d_code in target_depts:
        dept = db.query(models.DepartmentMaster).filter_by(department_code=d_code).first()
        if not dept:
            raise HTTPException(status_code=400, detail=f"Department {d_code} does not exist")

    is_honours = req.is_honours or ('H' in req.course_code[-3:].upper())
    
    created_count = 0
    for d_code in target_depts:
        existing = db.query(models.CourseMaster).filter_by(
            course_code=req.course_code, department_code=d_code
        ).first()
        
        if existing:
            if d_code == req.department_code:
                raise HTTPException(status_code=400, detail=f"Course {req.course_code} already exists in {d_code}")
            continue

        course = models.CourseMaster(
            course_code=req.course_code,
            course_name=req.course_name,
            department_code=d_code,
            semester=req.semester,
            lecture_hours=req.lecture_hours,
            tutorial_hours=req.tutorial_hours,
            practical_hours=req.practical_hours,
            credits=req.credits,
            weekly_sessions=req.weekly_sessions,
            is_lab=req.is_lab,
            is_honours=is_honours,
            is_minor=req.is_minor,
            is_elective=req.is_elective,
            is_open_elective=req.is_open_elective,
            is_add_course=req.is_add_course
        )
        db.add(course)
        created_count += 1

    if len(target_depts) > 1:
        db.query(models.CommonCourseMap).filter_by(
            course_code=req.course_code, semester=req.semester
        ).delete()
        for d_code in target_depts:
            db.add(models.CommonCourseMap(
                course_code=req.course_code,
                semester=req.semester,
                department_code=d_code
            ))

    db.commit()
    return {"status": "success", "course_code": req.course_code, "created_records": created_count}

@app.put("/courses/{code:path}")
def update_course(code: str, req: schemas.CourseUpdate, db: Session = Depends(get_db)):
    print(f"DEBUG PUT COURSE: received code={repr(code)}")
    # Find all course masters with this code
    existing_courses = db.query(models.CourseMaster).filter_by(course_code=code).all()
    print(f"DEBUG PUT COURSE: found existing: {existing_courses}")
    if not existing_courses:
        raise HTTPException(status_code=404, detail="Course not found")
        
    base_course = existing_courses[0]
    
    # 0. Handle course_code Primary Key Migration
    new_code = req.course_code if req.course_code is not None else code
    if new_code != code:
        # Check if new_code already exists
        conflict = db.query(models.CourseMaster).filter_by(course_code=new_code).first()
        if conflict:
            raise HTTPException(status_code=400, detail=f"Course {new_code} already exists")
            
        # 0.1 Create identical CourseMaster records mapped to new_code
        for existing in existing_courses:
            db.add(models.CourseMaster(
                course_code=new_code,
                department_code=existing.department_code,
                semester=existing.semester,
                course_name=existing.course_name,
                course_category=existing.course_category,
                delivery_type=existing.delivery_type,
                lecture_hours=existing.lecture_hours,
                tutorial_hours=existing.tutorial_hours,
                practical_hours=existing.practical_hours,
                credits=existing.credits,
                weekly_sessions=existing.weekly_sessions,
                is_lab=existing.is_lab,
                is_elective=existing.is_elective,
                is_open_elective=existing.is_open_elective,
                is_honours=existing.is_honours,
                is_minor=existing.is_minor,
                is_add_course=existing.is_add_course,
                enrolled_students=existing.enrolled_students
            ))
        db.flush() # Ensure DB sees the new masters before migrating children
        
        # 0.2 Migrate all Foreign Key constraint children to new_code
        db.query(models.CourseFacultyMap).filter_by(course_code=code).update({"course_code": new_code})
        db.query(models.CourseVenueMap).filter_by(course_code=code).update({"course_code": new_code})
        db.query(models.CommonCourseMap).filter_by(course_code=code).update({"course_code": new_code})
        db.query(models.CourseRegistration).filter_by(course_code=code).update({"course_code": new_code})
        db.query(models.TimetableEntry).filter_by(course_code=code).update({"course_code": new_code})
        
        # 0.3 Delete original CourseMaster orphans
        for existing in existing_courses:
            db.delete(existing)
        db.flush()
        
        # 0.4 Re-fetch and lock state to the new_code context
        existing_courses = db.query(models.CourseMaster).filter_by(course_code=new_code).all()
        base_course = existing_courses[0]
        code = new_code

    primary_dept = req.department_code if req.department_code is not None else base_course.department_code
    semester = req.semester if req.semester is not None else base_course.semester
    
    target_depts = [primary_dept]
    if hasattr(req, 'common_departments') and req.common_departments is not None:
        for d in req.common_departments:
            if d not in target_depts:
                target_depts.append(d)
    else:
        # If common_departments wasn't sent, keep existing mappings
        common_maps = db.query(models.CommonCourseMap).filter_by(course_code=code).all()
        for cm in common_maps:
            if cm.department_code not in target_depts:
                target_depts.append(cm.department_code)

    # Validate DEPARTMENTS
    for d_code in target_depts:
        if not db.query(models.DepartmentMaster).filter_by(department_code=d_code).first():
            raise HTTPException(status_code=400, detail=f"Department {d_code} does not exist")

    is_honours = req.is_honours if req.is_honours is not None else base_course.is_honours
    # Auto-detect honours if changed course code ends in H? (course code edit not supported via PUT right now)
    
    # 1. DELETE course masters that are no longer in target_depts
    for existing in existing_courses:
        if existing.department_code not in target_depts:
            db.delete(existing)
            
    # 2. UPDATE/CREATE target_depts
    for d_code in target_depts:
        course = db.query(models.CourseMaster).filter_by(course_code=code, department_code=d_code).first()
        if not course:
            course = models.CourseMaster(course_code=code, department_code=d_code)
            db.add(course)
            
        if req.course_name is not None: course.course_name = req.course_name
        if req.semester is not None: course.semester = req.semester
        if req.lecture_hours is not None: course.lecture_hours = req.lecture_hours
        if req.tutorial_hours is not None: course.tutorial_hours = req.tutorial_hours
        if req.practical_hours is not None: course.practical_hours = req.practical_hours
        if req.credits is not None: course.credits = req.credits
        if req.weekly_sessions is not None: course.weekly_sessions = req.weekly_sessions
        if req.is_lab is not None: course.is_lab = req.is_lab
        if req.is_honours is not None: course.is_honours = req.is_honours
        if req.is_minor is not None: course.is_minor = req.is_minor
        if req.is_elective is not None: course.is_elective = req.is_elective
        if req.is_open_elective is not None: course.is_open_elective = req.is_open_elective
        if req.is_add_course is not None: course.is_add_course = req.is_add_course

    # 3. Rebuild CommonCourseMap if more than 1 dept
    db.query(models.CommonCourseMap).filter_by(course_code=code).delete()
    if len(target_depts) > 1:
        for d_code in target_depts:
            db.add(models.CommonCourseMap(
                course_code=code,
                semester=semester,
                department_code=d_code
            ))
            
    db.commit()
    return {"status": "success", "course_code": code}

@app.delete("/courses/{code:path}")
def delete_course(code: str, db: Session = Depends(get_db)):
    courses = db.query(models.CourseMaster).filter_by(course_code=code).all()
    if not courses:
        raise HTTPException(status_code=404, detail="Course not found")
    
    # Remove course-faculty mappings
    db.query(models.CourseFacultyMap).filter_by(course_code=code).delete()
    # Remove common course mapping
    db.query(models.CommonCourseMap).filter_by(course_code=code).delete()
    
    for c in courses:
        db.delete(c)
    db.commit()
    return {"status": "deleted"}

# ============================================
# COURSE-FACULTY MAPPING
# ============================================

@app.get("/course-faculty")
def get_course_faculty(department_code: str = None, db: Session = Depends(get_db)):
    query = db.query(models.CourseFacultyMap)
    if department_code:
        query = query.filter(models.CourseFacultyMap.department_code == department_code)
    mappings = query.all()
    result = []
    for m in mappings:
        fac = db.query(models.FacultyMaster).filter_by(faculty_id=m.faculty_id).first()
        course = db.query(models.CourseMaster).filter_by(course_code=m.course_code).first()
        result.append({
            "id": m.id,
            "course_code": m.course_code,
            "course_name": course.course_name if course else m.course_code,
            "course_dept": course.department_code if course else "?",
            "course_semester": course.semester if course else 0,
            "faculty_id": m.faculty_id,
            "faculty_name": fac.faculty_name if fac else m.faculty_id,
            "faculty_dept": fac.department_code if fac else "?",
            "for_department": m.department_code,
            "delivery_type": m.delivery_type,
        })
    return result

@app.post("/course-faculty")
def create_course_faculty(req: schemas.CourseFacultyCreate, db: Session = Depends(get_db)):
    # Validate
    course = db.query(models.CourseMaster).filter_by(course_code=req.course_code).first()
    if not course:
        raise HTTPException(status_code=400, detail=f"Course {req.course_code} does not exist")
    fac = db.query(models.FacultyMaster).filter_by(faculty_id=req.faculty_id).first()
    if not fac:
        raise HTTPException(status_code=400, detail=f"Faculty {req.faculty_id} does not exist")
    # Check duplicate
    existing = db.query(models.CourseFacultyMap).filter_by(
        course_code=req.course_code, faculty_id=req.faculty_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="This mapping already exists")
    m = models.CourseFacultyMap(
        course_code=req.course_code,
        faculty_id=req.faculty_id,
        department_code=req.department_code,
        delivery_type=req.delivery_type
    )
    db.add(m)
    db.commit()
    return {"status": "success", "id": m.id}

@app.delete("/course-faculty/{mid}")
def delete_course_faculty(mid: int, db: Session = Depends(get_db)):
    m = db.query(models.CourseFacultyMap).filter_by(id=mid).first()
    if not m:
        raise HTTPException(status_code=404, detail="Mapping not found")
    db.delete(m)
    db.commit()
    return {"status": "deleted"}

# ============================================
# SLOTS
# ============================================

@app.get("/slots", response_model=List[schemas.Slot])
def get_slots(db: Session = Depends(get_db)):
    slots = db.query(models.SlotMaster).order_by(models.SlotMaster.day_of_week, models.SlotMaster.period_number).all()
    out = []
    for s in slots:
        try:
            parsed_ids = json.loads(s.semester_ids) if s.semester_ids else []
        except:
            parsed_ids = []
            
        out.append(schemas.Slot(
            slot_id=s.slot_id,
            day_of_week=s.day_of_week,
            period_number=s.period_number,
            start_time=s.start_time,
            end_time=s.end_time,
            slot_type=s.slot_type,
            is_active=s.is_active,
            semester_ids=parsed_ids
        ))
    return out

@app.post("/slots", response_model=schemas.Slot)
def create_slot(req: schemas.SlotCreate, db: Session = Depends(get_db)):
    existing = db.query(models.SlotMaster).filter_by(
        day_of_week=req.day_of_week, period_number=req.period_number
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Slot already exists for {req.day_of_week} Period {req.period_number}")
    slot = models.SlotMaster(
        day_of_week=req.day_of_week,
        period_number=req.period_number,
        start_time=req.start_time,
        end_time=req.end_time,
        slot_type=req.slot_type,
        is_active=req.is_active,
        semester_ids=json.dumps(req.semester_ids) if req.semester_ids is not None else "[]"
    )
    db.add(slot)
    db.commit()
    db.refresh(slot)
    
    try: parsed_ids = json.loads(slot.semester_ids)
    except: parsed_ids = []
    
    return schemas.Slot(
        slot_id=slot.slot_id,
        day_of_week=slot.day_of_week,
        period_number=slot.period_number,
        start_time=slot.start_time,
        end_time=slot.end_time,
        slot_type=slot.slot_type,
        is_active=slot.is_active,
        semester_ids=parsed_ids
    )

@app.put("/slots/{slot_id}", response_model=schemas.Slot)
def update_slot(slot_id: int, req: schemas.SlotUpdate, db: Session = Depends(get_db)):
    slot = db.query(models.SlotMaster).filter_by(slot_id=slot_id).first()
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    if req.start_time is not None:
        slot.start_time = req.start_time
    if req.end_time is not None:
        slot.end_time = req.end_time
    if req.slot_type is not None:
        slot.slot_type = req.slot_type
    if req.is_active is not None:
        slot.is_active = req.is_active
    if req.semester_ids is not None:
        slot.semester_ids = json.dumps(req.semester_ids)
        
    db.commit()
    db.refresh(slot)
    
    try: parsed_ids = json.loads(slot.semester_ids)
    except: parsed_ids = []
    
    return schemas.Slot(
        slot_id=slot.slot_id,
        day_of_week=slot.day_of_week,
        period_number=slot.period_number,
        start_time=slot.start_time,
        end_time=slot.end_time,
        slot_type=slot.slot_type,
        is_active=slot.is_active,
        semester_ids=parsed_ids
    )

@app.delete("/slots/{slot_id}")
def delete_slot(slot_id: int, db: Session = Depends(get_db)):
    slot = db.query(models.SlotMaster).filter_by(slot_id=slot_id).first()
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    db.delete(slot)
    db.commit()
    return {"status": "deleted", "slot_id": slot_id}

# ============================================
# BREAK CONFIGS
# ============================================

@app.get("/breaks", response_model=List[schemas.BreakConfig])
def get_breaks(db: Session = Depends(get_db)):
    breaks = db.query(models.BreakConfigMaster).all()
    out = []
    for b in breaks:
        try:
            parsed_ids = json.loads(b.semester_ids) if b.semester_ids else []
        except:
            parsed_ids = []
            
        out.append(schemas.BreakConfig(
            id=b.id,
            break_type=b.break_type,
            start_time=b.start_time,
            end_time=b.end_time,
            semester_ids=parsed_ids
        ))
    return out

@app.post("/breaks", response_model=schemas.BreakConfig)
def create_break(req: schemas.BreakConfigCreate, db: Session = Depends(get_db)):
    b = models.BreakConfigMaster(
        break_type=req.break_type,
        start_time=req.start_time,
        end_time=req.end_time,
        semester_ids=json.dumps(req.semester_ids) if req.semester_ids is not None else "[]"
    )
    db.add(b)
    db.commit()
    db.refresh(b)
    
    try: parsed_ids = json.loads(b.semester_ids)
    except: parsed_ids = []
    
    return schemas.BreakConfig(
        id=b.id,
        break_type=b.break_type,
        start_time=b.start_time,
        end_time=b.end_time,
        semester_ids=parsed_ids
    )

@app.put("/breaks/{break_id}", response_model=schemas.BreakConfig)
def update_break(break_id: int, req: schemas.BreakConfigUpdate, db: Session = Depends(get_db)):
    b = db.query(models.BreakConfigMaster).filter_by(id=break_id).first()
    if not b:
        raise HTTPException(status_code=404, detail="Break config not found")
        
    if req.break_type is not None:
        b.break_type = req.break_type
    if req.start_time is not None:
        b.start_time = req.start_time
    if req.end_time is not None:
        b.end_time = req.end_time
    if req.semester_ids is not None:
        b.semester_ids = json.dumps(req.semester_ids)
        
    db.commit()
    db.refresh(b)
    
    try: parsed_ids = json.loads(b.semester_ids)
    except: parsed_ids = []
    
    return schemas.BreakConfig(
        id=b.id,
        break_type=b.break_type,
        start_time=b.start_time,
        end_time=b.end_time,
        semester_ids=parsed_ids
    )

@app.delete("/breaks/{break_id}")
def delete_break(break_id: int, db: Session = Depends(get_db)):
    b = db.query(models.BreakConfigMaster).filter_by(id=break_id).first()
    if not b:
        raise HTTPException(status_code=404, detail="Break config not found")
    db.delete(b)
    db.commit()
    return {"status": "deleted", "id": break_id}

# ============================================
# GENERATION & TIMETABLE
# ============================================

@app.post("/generate")
def generate_timetable(request: schemas.GenerateRequest, db: Session = Depends(get_db)):
    from solver_engine import generate_schedule
    
    # 1. Fetch current config
    config_record = db.query(models.SchedulerConfig).first()
    
    config = DEFAULT_CONFIG
    if config_record and config_record.config_json:
        if isinstance(config_record.config_json, str):
            config = json.loads(config_record.config_json)
        else:
            config = config_record.config_json
            
    hard_mode = False
    try:
        hard_mode_cfg = config.get("validation", {}).get("hard_constraint_mode", {})
        hard_mode = bool(hard_mode_cfg.get("enabled", False))
    except (KeyError, AttributeError):
        pass
    
    result = generate_schedule(
        db, 
        department_code=request.department_code, 
        semester=request.semester,
        mentor_day=request.mentor_day,
        mentor_period=request.mentor_period,
        hard_mode=hard_mode
    )
    
    # Check if the result is the new dictionary format or the old boolean format
    if isinstance(result, bool):
        success = result
        if success:
            return {"status": "success", "message": "Timetable generated successfully", "warnings": []}
        else:
            courses = db.query(models.CourseMaster).filter_by(
                department_code=request.department_code, semester=request.semester, is_open_elective=False
            ).all()
            if not courses:
                raise HTTPException(status_code=400, detail=f"No courses found for {request.department_code} Sem {request.semester}. Check data import.")
            total_ws = sum(c.weekly_sessions for c in courses)
            raise HTTPException(status_code=400, detail=f"Cannot generate: {total_ws} weekly sessions for ~47 available slots. Schedule may be overloaded. Review course data for {request.department_code} Sem {request.semester}.")
            
    # New structured response handling
    if not result.get("success", False):
        if result.get("errors"):
            # If we have structured errors and hard mode blocked generation, return 422 Unprocessable Entity
            raise HTTPException(status_code=422, detail={
                "message": "Resource verification failed. Schedule generation aborted.",
                "errors": result.get("errors", []),
                "warnings": result.get("warnings", [])
            })
        else:
            raise HTTPException(status_code=400, detail="Solver exhausted all possibilities and failed to find a valid arrangement. Try reducing constraints.")
            
    # Success with potential warnings
    return {
        "status": "success", 
        "message": f"Successfully generated and saved {result.get('entries_saved', 0)} entries.", 
        "warnings": result.get("warnings", []),
        "entries_saved": result.get("entries_saved", 0)
    }

# ============================================
# UPDATED TIMETABLE ENDPOINT (MASTER VIEW)
# ============================================
@app.get("/timetable", response_model=List[schemas.TimetableEntry])
def get_timetable(
    department_code: Optional[str] = None, 
    semester: Optional[int] = None, 
    db: Session = Depends(get_db)
):
    query = db.query(models.TimetableEntry)
    
    # Logic: Only apply filter if the parameter is NOT None AND NOT an empty string
    # This ignores empty strings sent by frontend (e.g. ?department_code="")
    if department_code and department_code.strip():
        query = query.filter(models.TimetableEntry.department_code == department_code)
        
    if semester:
        query = query.filter(models.TimetableEntry.semester == semester)
        
    # If no valid filters are provided, this returns ALL entries (Master View)
    return query.all()

@app.post("/timetable/save")
def save_timetable(request: schemas.TimetableSaveRequest, db: Session = Depends(get_db)):
    # Delete existing entries for this Dept/Semester
    db.query(models.TimetableEntry).filter_by(
        department_code=request.department_code,
        semester=request.semester
    ).delete()
    
    # Bulk insert new entries
    new_entries = [models.TimetableEntry(**entry.dict()) for entry in request.entries]
    db.add_all(new_entries)
    
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Save failed: {str(e)}")

@app.get("/api/conflicts")
def get_conflicts(
    department_code: Optional[str] = None, 
    semester: Optional[int] = None, 
    db: Session = Depends(get_db)
):
    from conflict_detector import detect_conflicts
    return detect_conflicts(db, department_code, semester)
        
    return {"status": "success", "count": len(new_entries)}

# ============================================
# AVAILABILITY QUERIES (for Smart Editor)
# ============================================

@app.get("/available-faculty")
def get_available_faculty(
    department_code: str,
    day: str,
    period: int,
    course_code: str = None,
    show_all: bool = False,
    db: Session = Depends(get_db)
):
    """Return faculty, applying context-aware filtering based on course mapping."""
    # All faculty_names busy at this exact slot (across ALL departments and semesters)
    busy_names = set()
    busy_entries = db.query(models.TimetableEntry.faculty_name).filter(
        models.TimetableEntry.day_of_week == day,
        models.TimetableEntry.period_number == period,
        models.TimetableEntry.faculty_name.isnot(None),
        models.TimetableEntry.faculty_name != '',
        models.TimetableEntry.faculty_name != 'Unassigned'
    ).all()
    for (name,) in busy_entries:
        busy_names.add(name)

    base_faculty = []
    
    if show_all:
        base_faculty = db.query(models.FacultyMaster).all()
    elif course_code:
        # Try to find specific mappings for this course in this department
        mapped = db.query(models.CourseFacultyMap).filter_by(
            department_code=department_code, course_code=course_code
        ).all()
        if mapped:
            mapped_ids = [m.faculty_id for m in mapped]
            base_faculty = db.query(models.FacultyMaster).filter(
                models.FacultyMaster.faculty_id.in_(mapped_ids)
            ).all()
            
    # Fallback to department if not found or no course_code provided
    if not show_all and not base_faculty:
        base_faculty = db.query(models.FacultyMaster).filter_by(
            department_code=department_code
        ).all()

    result = []
    for f in base_faculty:
        result.append({
            "faculty_id": f.faculty_id,
            "faculty_name": f.faculty_name,
            "department_code": f.department_code,
            "is_available": f.faculty_name not in busy_names
        })
    # Sort: available first, then by name
    result.sort(key=lambda x: (not x["is_available"], x["faculty_name"]))
    return result


@app.get("/available-venues")
def get_available_venues(
    department_code: str,
    semester: int,
    day: str,
    period: int,
    course_code: str = None,
    show_all: bool = False,
    db: Session = Depends(get_db)
):
    """Return venues, applying context-aware filtering based on global or local course mappings."""
    # All venue_names busy at this exact slot (across ALL departments and semesters)
    busy_venue_names = set()
    busy_entries = db.query(models.TimetableEntry.venue_name).filter(
        models.TimetableEntry.day_of_week == day,
        models.TimetableEntry.period_number == period,
        models.TimetableEntry.venue_name.isnot(None),
        models.TimetableEntry.venue_name != ''
    ).all()
    for (name,) in busy_entries:
        busy_venue_names.add(name)

    base_venues = []

    if show_all:
        base_venues = db.query(models.VenueMaster).all()
    elif course_code:
        # 1. Check if it's a common course with a global venue lock
        common = db.query(models.CommonCourseMap).filter_by(
            course_code=course_code, semester=semester
        ).filter(models.CommonCourseMap.venue_name != None).first()
        
        if common and common.venue_name:
            v_obj = db.query(models.VenueMaster).filter_by(venue_name=common.venue_name).first()
            if v_obj: base_venues = [v_obj]
            
        # 2. If no global lock, check for department-specific mappings
        if not base_venues:
            mapped = db.query(models.CourseVenueMap).filter_by(
                department_code=department_code, course_code=course_code
            ).all()
            if mapped:
                mapped_ids = [m.venue_id for m in mapped]
                base_venues = db.query(models.VenueMaster).filter(
                    models.VenueMaster.venue_id.in_(mapped_ids)
                ).all()

    # Fallback to department venues if not found or no course_code provided
    if not show_all and not base_venues:
        dept_venues = db.query(models.DepartmentVenueMap).filter_by(
            department_code=department_code, semester=semester
        ).all()
        if dept_venues:
            mapped_vids = [v.venue_id for v in dept_venues]
            base_venues = db.query(models.VenueMaster).filter(
                models.VenueMaster.venue_id.in_(mapped_vids)
            ).all()
        else:
            # Absolute fallback if department has no venues mapped at all
            base_venues = db.query(models.VenueMaster).all()

    result = []
    for v in base_venues:
        result.append({
            "venue_id": v.venue_id,
            "venue_name": v.venue_name,
            "block": v.block,
            "is_lab": v.is_lab,
            "capacity": v.capacity,
            "is_available": v.venue_name not in busy_venue_names
        })
    # Sort: available first, then by name
    result.sort(key=lambda x: (not x["is_available"], x["venue_name"]))
    return result


# ============================================
# VENUES
# ============================================

@app.get("/venues", response_model=List[schemas.Venue])
def get_venues(db: Session = Depends(get_db)):
    return db.query(models.VenueMaster).all()

@app.post("/venues", response_model=schemas.Venue)
def create_venue(req: schemas.VenueCreate, db: Session = Depends(get_db)):
    # Check for duplicate
    existing = db.query(models.VenueMaster).filter_by(venue_name=req.venue_name).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Venue '{req.venue_name}' already exists")
    
    venue = models.VenueMaster(
        venue_name=req.venue_name,
        block=req.block,
        is_lab=req.is_lab,
        capacity=req.capacity
    )
    db.add(venue)
    try:
        db.commit()
        db.refresh(venue)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    return venue

@app.delete("/venues/{venue_id}")
def delete_venue(venue_id: int, db: Session = Depends(get_db)):
    venue = db.query(models.VenueMaster).filter_by(venue_id=venue_id).first()
    if not venue:
        raise HTTPException(status_code=404, detail="Venue not found")
    db.delete(venue)
    db.commit()
    return {"status": "deleted", "venue_id": venue_id}

@app.post("/venues/import")
def import_venues(db: Session = Depends(get_db)):
    import pandas as pd
    import os
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(base_dir)
    file_path = os.path.join(project_root, "data", "campus_classrooms_labs_simplified.xlsx")
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"File not found: campus_classrooms_labs_simplified.xlsx")
    
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    # Expected columns: 'Room Name', 'Block ID'
    if 'Room Name' not in df.columns:
        raise HTTPException(status_code=400, detail=f"Excel file missing 'Room Name' column. Found: {list(df.columns)}")
    
    imported_count = 0
    skipped_count = 0
    lab_keywords = ['lab', 'workshop', 'studio', 'drawing']
    
    for _, row in df.iterrows():
        venue_name = str(row.get('Room Name', '')).strip()
        block = str(row.get('Block ID', '')).strip() if pd.notna(row.get('Block ID')) else None
        
        if not venue_name or venue_name.lower() == 'nan':
            continue
        
        # Auto-detect if it's a lab based on name
        is_lab = any(kw in venue_name.lower() for kw in lab_keywords)
        
        # Skip if already exists
        existing = db.query(models.VenueMaster).filter_by(venue_name=venue_name).first()
        if existing:
            skipped_count += 1
            continue
        
        venue = models.VenueMaster(
            venue_name=venue_name,
            block=block,
            is_lab=is_lab,
            capacity=60
        )
        db.add(venue)
        imported_count += 1
    
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")
    
    return {"status": "success", "imported_count": imported_count, "skipped_count": skipped_count}

# --- Department Venue Mapping ---
@app.get("/department-venues", response_model=List[schemas.DepartmentVenueResponse])
def get_department_venues(department_code: str = None, semester: int = None, db: Session = Depends(get_db)):
    query = db.query(models.DepartmentVenueMap)
    if department_code:
        query = query.filter(models.DepartmentVenueMap.department_code == department_code)
    if semester:
        query = query.filter(models.DepartmentVenueMap.semester == semester)
    
    maps = query.all()
    result = []
    for m in maps:
        venue = db.query(models.VenueMaster).filter_by(venue_id=m.venue_id).first()
        if venue:
            result.append(schemas.DepartmentVenueResponse(
                id=m.id,
                department_code=m.department_code,
                semester=m.semester,
                venue_id=venue.venue_id,
                venue_name=venue.venue_name,
                is_lab=venue.is_lab,
                capacity=venue.capacity
            ))
    return result

@app.post("/department-venues", response_model=schemas.DepartmentVenueResponse)
def create_department_venue(req: schemas.DepartmentVenueCreate, db: Session = Depends(get_db)):
    existing = db.query(models.DepartmentVenueMap).filter_by(
        department_code=req.department_code,
        semester=req.semester,
        venue_id=req.venue_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="This venue is already mapped to this department's semester.")
    
    new_map = models.DepartmentVenueMap(department_code=req.department_code, semester=req.semester, venue_id=req.venue_id)
    db.add(new_map)
    db.commit()
    db.refresh(new_map)
    
    venue = db.query(models.VenueMaster).filter_by(venue_id=req.venue_id).first()
    return schemas.DepartmentVenueResponse(
        id=new_map.id,
        department_code=new_map.department_code,
        semester=new_map.semester,
        venue_id=venue.venue_id,
        venue_name=venue.venue_name,
        is_lab=venue.is_lab,
        capacity=venue.capacity
    )

@app.delete("/department-venues/{map_id}")
def delete_department_venue(map_id: int, db: Session = Depends(get_db)):
    m = db.query(models.DepartmentVenueMap).filter_by(id=map_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Mapping not found")
    db.delete(m)
    db.commit()
    return {"status": "success", "id": map_id}

# --- Course Venue Mapping ---
@app.get("/course-venues", response_model=List[schemas.CourseVenueResponse])
def get_course_venues(department_code: str = None, db: Session = Depends(get_db)):
    query = db.query(models.CourseVenueMap)
    if department_code:
        query = query.filter(models.CourseVenueMap.department_code == department_code)
    
    maps = query.all()
    result = []
    for m in maps:
        venue = db.query(models.VenueMaster).filter_by(venue_id=m.venue_id).first()
        if venue:
            result.append(schemas.CourseVenueResponse(
                id=m.id,
                department_code=m.department_code,
                course_code=m.course_code,
                venue_id=venue.venue_id,
                venue_name=venue.venue_name,
                is_lab=venue.is_lab,
                capacity=venue.capacity,
                venue_type=m.venue_type or 'BOTH'
            ))
    return result

@app.post("/course-venues", response_model=schemas.CourseVenueResponse)
def create_course_venue(req: schemas.CourseVenueCreate, db: Session = Depends(get_db)):
    # Block venue assignment for common courses — must use /common-courses/venue instead
    is_common = db.query(models.CommonCourseMap).filter_by(
        course_code=req.course_code
    ).first()
    if is_common:
        raise HTTPException(
            status_code=400,
            detail=f"'{req.course_code}' is a common course shared across departments. "
                   f"Use the Common Course Venues section to assign its venue globally."
        )

    existing = db.query(models.CourseVenueMap).filter_by(
        department_code=req.department_code,
        course_code=req.course_code,
        venue_id=req.venue_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="This venue is already mapped to the course.")
    
    new_map = models.CourseVenueMap(
        department_code=req.department_code, 
        course_code=req.course_code, 
        venue_id=req.venue_id,
        venue_type=(req.venue_type or 'BOTH').upper()
    )
    db.add(new_map)
    db.commit()
    db.refresh(new_map)
    
    venue = db.query(models.VenueMaster).filter_by(venue_id=req.venue_id).first()
    return schemas.CourseVenueResponse(
        id=new_map.id,
        department_code=new_map.department_code,
        course_code=new_map.course_code,
        venue_id=venue.venue_id,
        venue_name=venue.venue_name,
        is_lab=venue.is_lab,
        capacity=venue.capacity,
        venue_type=new_map.venue_type or 'BOTH'
    )

@app.delete("/course-venues/{map_id}")
def delete_course_venue(map_id: int, db: Session = Depends(get_db)):
    m = db.query(models.CourseVenueMap).filter_by(id=map_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Mapping not found")
    db.delete(m)
    db.commit()
    return {"status": "success", "id": map_id}

# ============================================
# STUDENTS & REGISTRATIONS
# ============================================

@app.get("/students", response_model=List[schemas.StudentResponse])
def get_students(department_code: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(models.StudentMaster)
    if department_code:
        query = query.filter_by(department_code=department_code)
    return query.all()

@app.get("/registrations", response_model=List[schemas.CourseRegistrationResponse])
def get_registrations(course_code: Optional[str] = None, semester: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(models.CourseRegistration)
    if course_code:
        query = query.filter_by(course_code=course_code)
    if semester:
        query = query.filter_by(semester=semester)
    return query.all()

@app.post("/registrations", response_model=schemas.CourseRegistrationResponse)
def create_registration(req: schemas.CourseRegistrationCreate, db: Session = Depends(get_db)):
    student = db.query(models.StudentMaster).filter_by(student_id=req.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    course = db.query(models.CourseMaster).filter_by(
        course_code=req.course_code, 
        semester=req.semester,
        department_code=student.department_code
    ).first()
    
    if not course:
        raise HTTPException(status_code=404, detail=f"Course {req.course_code} not mapped to this student's department ({student.department_code})")
        
    existing = db.query(models.CourseRegistration).filter_by(student_id=req.student_id, course_code=req.course_code, semester=req.semester).first()
    if existing:
        raise HTTPException(status_code=400, detail="Student already registered for this course")
        
    reg = models.CourseRegistration(**req.dict())
    db.add(reg)
    course.enrolled_students = (course.enrolled_students or 0) + 1
    db.commit()
    db.refresh(reg)
    return reg

@app.delete("/registrations/{reg_id}")
def delete_registration(reg_id: int, db: Session = Depends(get_db)):
    reg = db.query(models.CourseRegistration).filter_by(id=reg_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
        
    student = db.query(models.StudentMaster).filter_by(student_id=reg.student_id).first()
    if student:
        course = db.query(models.CourseMaster).filter_by(
            course_code=reg.course_code, 
            semester=reg.semester,
            department_code=student.department_code
        ).first()
        if course and course.enrolled_students and course.enrolled_students > 0:
            course.enrolled_students -= 1
        
    db.delete(reg)
    db.commit()
    return {"status": "success", "id": reg_id}

@app.post("/students", response_model=schemas.StudentResponse)
def create_student(req: schemas.StudentCreate, db: Session = Depends(get_db)):
    student = db.query(models.StudentMaster).filter_by(student_id=req.student_id).first()
    if student:
        raise HTTPException(status_code=400, detail="Student ID already exists")
    student = models.StudentMaster(**req.dict())
    db.add(student)
    db.commit()
    db.refresh(student)
    return student

@app.delete("/students/{student_id}")
def delete_student(student_id: str, db: Session = Depends(get_db)):
    student = db.query(models.StudentMaster).filter_by(student_id=student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
        
    regs = db.query(models.CourseRegistration).filter_by(student_id=student_id).all()
    for r in regs:
        course = db.query(models.CourseMaster).filter_by(
            course_code=r.course_code, 
            semester=r.semester,
            department_code=student.department_code
        ).first()
        if course and course.enrolled_students and course.enrolled_students > 0:
            course.enrolled_students -= 1
        db.delete(r)
        
    db.delete(student)
    db.commit()
    return {"status": "success", "student_id": student_id}

# ============================================
# PERSONALIZED TIMETABLES
# ============================================

@app.get("/timetable/faculty/{faculty_id}", response_model=schemas.PersonalTimetableResponse)
def get_faculty_timetable(faculty_id: str, db: Session = Depends(get_db)):
    entries = db.query(models.TimetableEntry).filter_by(faculty_id=faculty_id).all()
    
    slot_map = {}
    conflicts = set()
    for e in entries:
        key = (e.day_of_week, e.period_number)
        if key not in slot_map:
            slot_map[key] = [e]
        else:
            slot_map[key].append(e)

    for key, items in slot_map.items():
        if len(items) > 1:
            venues = set(i.venue_name for i in items if i.venue_name)
            if len(venues) > 1:
                courses_str = " & ".join([f"{i.course_code} ({i.venue_name})" for i in items])
                conflicts.add(f"Conflict on {key[0]} Period {key[1]}: Scheduled simultaneously in different venues across {courses_str}")

    return {"conflicts": list(conflicts), "timetable": entries}


@app.get("/timetable/student/{student_id}", response_model=schemas.PersonalTimetableResponse)
def get_student_timetable(student_id: str, db: Session = Depends(get_db)):
    student = db.query(models.StudentMaster).filter_by(student_id=student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
        
    regs = db.query(models.CourseRegistration).filter_by(student_id=student_id).all()
    
    valid_entries = {}
    student_semester = None
    
    # Pre-fetch Department master timetable to map Elective 5 slots for Open Electives
    dept_master_slots = []
    
    for r in regs:
        if student_semester is None:
            student_semester = r.semester
            # Cache the department's master timetable layout to find the Elective 5 slots
            dept_master_slots = db.query(models.TimetableEntry).filter_by(
                department_code=student.department_code, semester=student_semester
            ).all()

        course = db.query(models.CourseMaster).filter_by(course_code=r.course_code, semester=r.semester).first()
        is_open_elective = course and course.is_open_elective
        
        c_entries = db.query(models.TimetableEntry).filter_by(course_code=r.course_code, semester=r.semester).all()
        
        if is_open_elective and not c_entries:
            # If it's an Open Elective without explicit cross-department scheduling,
            # force it into the slots where the student's home department runs "ELECTIVE 5"
            elective_5_slots = []
            for dme in dept_master_slots:
                # Find the master Elective 5 block (or any block that is part of the Elective V group)
                # Since we don't strictly have course categories on TimetableEntry, we inspect the course
                dme_course = db.query(models.CourseMaster).filter_by(course_code=dme.course_code, semester=r.semester).first()
                if dme_course and dme_course.is_elective and (dme_course.course_category == "ELECTIVE 5" or dme_course.course_category == "ELECTIVE V"):
                    # Include slot_id to satisfy Pydantic schema constraints
                    elective_5_slots.append((dme.day_of_week, dme.period_number, dme.slot_id))
            
            # Deduplicate the coordinates in case multiple Elective 5s share the slot
            elective_5_coords = list(set(elective_5_slots))
            
            for (day, period, slot_id) in elective_5_coords:
                # Mock a TimetableEntry dynamically for the render engine
                mock_entry = models.TimetableEntry(
                    id=999900 + len(valid_entries),  # Dummy ID
                    course_code=r.course_code,
                    course_name=course.course_name,
                    department_code=student.department_code,
                    semester=r.semester,
                    day_of_week=day,
                    period_number=period,
                    session_type="THEORY",
                    faculty_name="Unassigned",
                    venue_name="",
                    section_number=1,
                    slot_id=slot_id
                )
                k = (r.course_code, day, period)
                valid_entries[k] = mock_entry

        else:
            for e in c_entries:
                # Display section 1 as the default schedule for students
                if e.section_number == 1 or e.section_number is None:
                    k = (e.course_code, e.day_of_week, e.period_number)
                    if k not in valid_entries:
                        valid_entries[k] = e
                    
    # Also fetch common courses that are natively meant for this student's department/semester
    # but don't explicitly require personal registration (e.g. Mentor Hour, Mini Project)
    if student_semester is not None:
        common_courses = ['mentor', 'mini']
        for e in dept_master_slots:
            if e.course_name and e.course_code and any(c in e.course_name.lower() or c in e.course_code.lower() for c in common_courses):
                if e.section_number == 1 or e.section_number is None:
                    k = (e.course_code, e.day_of_week, e.period_number)
                    if k not in valid_entries:
                        valid_entries[k] = e

    # --- Mini Project Fallback: Fill Add Course slots with "Mini Project" for non-opted students ---
    if student_semester is not None:
        dept_rec = db.query(models.DepartmentMaster).filter_by(department_code=student.department_code).first()
        if dept_rec and dept_rec.pair_add_course_miniproject:
            # Find all Add Course subjects for this department/semester
            add_courses = db.query(models.CourseMaster).filter_by(
                department_code=student.department_code, semester=student_semester, is_add_course=True
            ).all()
            # Get the student's registered course codes
            student_course_codes = set(r.course_code for r in regs)
            
            for ac in add_courses:
                if ac.course_code not in student_course_codes:
                    # Student did NOT opt for this Add Course — find its scheduled slots
                    ac_entries = db.query(models.TimetableEntry).filter_by(
                        course_code=ac.course_code, semester=student_semester
                    ).all()
                    for ace in ac_entries:
                        if ace.section_number == 1 or ace.section_number is None:
                            k = ('MINI_PROJECT_FALLBACK', ace.day_of_week, ace.period_number)
                            if k not in valid_entries:
                                mock_mp = models.TimetableEntry(
                                    id=999800 + len(valid_entries),
                                    course_code='MINI_PROJECT',
                                    course_name='Mini Project',
                                    department_code=student.department_code,
                                    semester=student_semester,
                                    day_of_week=ace.day_of_week,
                                    period_number=ace.period_number,
                                    session_type='THEORY',
                                    faculty_name='Unassigned',
                                    venue_name='',
                                    section_number=1,
                                    slot_id=ace.slot_id
                                )
                                valid_entries[k] = mock_mp

    entries = list(valid_entries.values())
    
    import re
    # Hydrate entries with course classification flags for the frontend labels,
    # and clean up noisy course names (like `/ OPEN ELECTIVE`) for regular electives.
    for e in entries:
        course = db.query(models.CourseMaster).filter_by(course_code=e.course_code, semester=e.semester).first()
        if course:
            setattr(e, 'is_honours', course.is_honours)
            setattr(e, 'is_minor', course.is_minor)
            setattr(e, 'is_elective', course.is_elective)
            setattr(e, 'is_open_elective', course.is_open_elective)
            setattr(e, 'is_add_course', course.is_add_course)
            
            # Clean up inherited UI noise in the course name (for regular electives)
            if course.is_elective and not course.is_open_elective and e.course_name:
                e.course_name = re.sub(r'(?i)\s*/\s*OPEN\s*ELECTIVE\s*', '', e.course_name).strip()
                e.course_name = re.sub(r'(?i)\s*-\s*OPEN\s*ELECTIVE\s*', '', e.course_name).strip()
        else:
            setattr(e, 'is_honours', False)
            setattr(e, 'is_minor', False)
            setattr(e, 'is_elective', False)
            setattr(e, 'is_open_elective', False)
            setattr(e, 'is_add_course', False)
    
    slot_map = {}
    conflicts = set()
    for e in entries:
        key = (e.day_of_week, e.period_number)
        if key not in slot_map:
            slot_map[key] = [e]
        else:
            slot_map[key].append(e)
            
    for key, items in slot_map.items():
        if len(items) > 1:
            course_codes = set(i.course_code for i in items)
            if len(course_codes) > 1:
                courses_str = " & ".join(course_codes)
                conflicts.add(f"Conflict on {key[0]} Period {key[1]}: Student is registered for colliding courses: {courses_str}")

    return {"conflicts": list(conflicts), "timetable": entries}


@app.get("/timetable/venue/{venue_name}", response_model=schemas.PersonalTimetableResponse)
def get_venue_timetable(venue_name: str, db: Session = Depends(get_db)):
    # 1. Fetch matching entries (can be part of comma separated venues)
    all_entries = db.query(models.TimetableEntry).all()
    raw_entries = []
    for e in all_entries:
        if e.venue_name and venue_name in [v.strip() for v in e.venue_name.split(",")]:
            raw_entries.append(e)

    # 2. Build set of common course codes for fast lookup
    all_common = db.query(models.CommonCourseMap).all()
    common_course_keys = set()  # {(course_code, semester)}
    for cc in all_common:
        common_course_keys.add((cc.course_code, cc.semester))

    # 3. Group entries by (day, period) to detect common courses vs conflicts
    slot_groups = {}  # (day, period) -> [entry, ...]
    for e in raw_entries:
        key = (e.day_of_week, e.period_number)
        slot_groups.setdefault(key, []).append(e)

    # 4. Process each slot: compute strength, detect common courses, merge breakdown
    entries = []
    conflicts = set()

    for slot_key, slot_entries in slot_groups.items():
        # Sub-group by course_code within the same slot
        course_groups = {}  # course_code -> [entries for that course]
        for e in slot_entries:
            course_groups.setdefault(e.course_code, []).append(e)

        # Check for REAL conflicts: different course_codes in same venue+slot
        if len(course_groups) > 1:
            courses_str = " & ".join(
                [f"{code} ({', '.join(set(x.department_code for x in grp))})"
                 for code, grp in course_groups.items()]
            )
            conflicts.add(
                f"Conflict on {slot_key[0]} Period {slot_key[1]}: "
                f"Venue occupied by multiple distinct courses: {courses_str}"
            )

        # Process each course group in this slot
        for course_code, group in course_groups.items():
            is_common = (course_code, group[0].semester) in common_course_keys and len(group) > 1

            # Build department breakdown with strength per dept
            dept_breakdown = []
            total_combined = 0

            for e in group:
                # Calculate per-dept strength
                sem_count = db.query(models.DepartmentSemesterCount).filter_by(
                    department_code=e.department_code, semester=e.semester
                ).first()
                student_count = sem_count.student_count if sem_count and sem_count.student_count > 0 else 60

                course = db.query(models.CourseMaster).filter_by(
                    course_code=e.course_code, semester=e.semester, department_code=e.department_code
                ).first()
                enrolled = course.enrolled_students if course and course.enrolled_students else 0
                base_count = min(enrolled, student_count) if enrolled > 0 else student_count

                # Calculate sections for this dept's version of the course
                same_course_entries = db.query(models.TimetableEntry).filter_by(
                    course_code=e.course_code, semester=e.semester,
                    department_code=e.department_code, session_type=e.session_type
                ).all()
                total_secs = max([se.section_number or 1 for se in same_course_entries] + [1])
                strength = math.ceil(base_count / total_secs) if total_secs > 0 else base_count

                dept_breakdown.append({
                    "dept": e.department_code,
                    "count": strength,
                    "total_sections": total_secs,
                    "section": e.section_number or 1
                })
                total_combined += strength

                setattr(e, 'total_sections', total_secs)
                setattr(e, 'strength', strength)

            if is_common:
                # For common courses, attach breakdown and combined strength to ALL entries in group
                for e in group:
                    setattr(e, 'dept_breakdown', dept_breakdown)
                    setattr(e, 'combined_strength', total_combined)
                    setattr(e, 'is_common_course', True)

                # Only emit the FIRST entry to avoid duplicate rows in the timetable grid
                # but with complete aggregated data
                representative = group[0]
                setattr(representative, 'dept_breakdown', dept_breakdown)
                setattr(representative, 'combined_strength', total_combined)
                setattr(representative, 'is_common_course', True)
                entries.append(representative)
            else:
                # Non-common: add all entries individually (could be split batches, etc.)
                for e in group:
                    setattr(e, 'dept_breakdown', None)
                    setattr(e, 'combined_strength', None)
                    setattr(e, 'is_common_course', False)
                    entries.append(e)

    return {"conflicts": list(conflicts), "timetable": entries}

# ============================================
# CONFLICT CHECK (Cross-Department)
# ============================================

@app.post("/check-conflicts")
def check_conflicts(request: schemas.ConflictCheckRequest, db: Session = Depends(get_db)):
    """
    Check if any faculty or venue in the given entries is already occupied
    by another department/semester at the same day+period.
    Returns a list of conflict details.
    """
    faculty_conflicts = []
    venue_conflicts = []
    
    for entry in request.entries:
        day = entry.day_of_week
        period = entry.period_number
        
        # Check faculty conflicts
        if entry.faculty_name and entry.faculty_name.strip() and entry.faculty_name not in ('', 'Unassigned'):
            busy = db.query(models.TimetableEntry).filter(
                models.TimetableEntry.day_of_week == day,
                models.TimetableEntry.period_number == period,
                models.TimetableEntry.faculty_name == entry.faculty_name,
                # Exclude entries from the SAME dept+sem (those are being replaced)
                ~(
                    (models.TimetableEntry.department_code == request.department_code) &
                    (models.TimetableEntry.semester == request.semester)
                )
            ).all()
            
            for b in busy:
                conflict_msg = f"{entry.faculty_name} is occupied by {b.department_code} Sem {b.semester} ({b.course_code}) at {day} Period {period}"
                if conflict_msg not in [c['message'] for c in faculty_conflicts]:
                    faculty_conflicts.append({
                        "type": "faculty",
                        "name": entry.faculty_name,
                        "day": day,
                        "period": period,
                        "occupied_by_dept": b.department_code,
                        "occupied_by_sem": b.semester,
                        "occupied_by_course": b.course_code,
                        "message": conflict_msg
                    })
        
        # Check venue conflicts
        if entry.venue_name and entry.venue_name.strip():
            busy = db.query(models.TimetableEntry).filter(
                models.TimetableEntry.day_of_week == day,
                models.TimetableEntry.period_number == period,
                models.TimetableEntry.venue_name == entry.venue_name,
                ~(
                    (models.TimetableEntry.department_code == request.department_code) &
                    (models.TimetableEntry.semester == request.semester)
                )
            ).all()
            
            for b in busy:
                conflict_msg = f"{entry.venue_name} is occupied by {b.department_code} Sem {b.semester} ({b.course_code}) at {day} Period {period}"
                if conflict_msg not in [c['message'] for c in venue_conflicts]:
                    venue_conflicts.append({
                        "type": "venue",
                        "name": entry.venue_name,
                        "day": day,
                        "period": period,
                        "occupied_by_dept": b.department_code,
                        "occupied_by_sem": b.semester,
                        "occupied_by_course": b.course_code,
                        "message": conflict_msg
                    })
    
    return {
        "has_conflicts": len(faculty_conflicts) > 0 or len(venue_conflicts) > 0,
        "faculty_conflicts": faculty_conflicts,
        "venue_conflicts": venue_conflicts
    }


# ============================================
# SCHEDULER CONFIG (Constraint Management)
# ============================================
# ============================================
import json

DEFAULT_CONFIG = {
    "validation": {
        "hard_constraint_mode": {
            "value": False, "enabled": False, "type": "flag",
            "label": "Hard Constraint Mode",
            "description": "Block generation and display errors when there are insufficient faculty or venues"
        }
    },
    "hard_constraints": {
        "max_courses_per_slot": {
            "value": 1, "enabled": True, "type": "number",
            "label": "Max Courses Per Slot",
            "description": "Maximum number of different courses allowed in a single time slot"
        },
        "lab_block_starts": {
            "value": [1, 3, 5], "enabled": True, "type": "array",
            "label": "Lab Block Start Periods",
            "description": "Periods where 2-period lab blocks can begin (e.g., P1-P2, P3-P4, P5-P6)"
        },
        "max_lab_blocks_per_day": {
            "value": 1, "enabled": True, "type": "number",
            "label": "Max Lab Blocks Per Day",
            "description": "Maximum number of lab blocks (2-period) allowed across ALL courses in a single day"
        },
        "mentor_hour_blocked": {
            "value": True, "enabled": True, "type": "boolean",
            "label": "Block Mentor Hour",
            "description": "Prevent any regular course from being scheduled during the mentor interaction hour"
        },
        "no_faculty_clash": {
            "value": True, "enabled": True, "type": "boolean",
            "label": "No Faculty Clash",
            "description": "Same faculty cannot teach two different courses simultaneously"
        },
        "no_theory_in_own_lab": {
            "value": True, "enabled": True, "type": "boolean",
            "label": "No Theory in Own Lab Block",
            "description": "A course cannot have theory in the same period as its own lab block"
        }
    },
    "dynamic_constraints": {
        "max_theory_per_course_per_day": {
            "value": 1, "overloaded_value": 2, "enabled": True, "type": "number",
            "label": "Max Theory/Course/Day",
            "description": "Maximum theory periods for the same course in a single day (normal mode)"
        },
        "max_theory_per_course_per_day_overloaded": {
            "value": 2, "enabled": True, "type": "number",
            "label": "Max Theory/Course/Day (Overloaded)",
            "description": "Maximum theory periods per course per day when schedule is overloaded"
        },
        "no_back_to_back_theory": {
            "value": True, "enabled": True, "type": "boolean",
            "label": "No Back-to-Back Theory",
            "description": "Prevent consecutive theory periods of the same course (relaxed when overloaded)"
        },
        "p8_honours_only": {
            "value": True, "enabled": True, "type": "boolean",
            "label": "Period 8 for Honours Only",
            "description": "Reserve Period 8 exclusively for Honours/Minor courses (opens for regular if no honours)"
        }
    },
    "soft_constraints": {
        "non_consecutive_lab_days_penalty": {
            "value": -5, "enabled": True, "type": "number",
            "label": "Consecutive Lab Days Penalty",
            "description": "Penalty applied when lab blocks are scheduled on consecutive days (negative = penalty)"
        },
        "theory_lab_same_day_bonus": {
            "value": 3, "enabled": True, "type": "number",
            "label": "Theory+Lab Same Day Bonus",
            "description": "Bonus for scheduling theory on the same day as its lab session"
        },
        "fill_slots_bonus": {
            "value": 10, "enabled": True, "type": "number",
            "label": "Filled Slot Bonus",
            "description": "Bonus per slot that gets filled (maximizes slot utilization)"
        }
    },
    "section_settings": {
        "min_section_threshold": {
            "value": 30, "enabled": True, "type": "number",
            "label": "Min Students for Extra Section",
            "description": "Minimum number of remaining students needed to create an additional section"
        },
        "default_venue_capacity": {
            "value": 60, "enabled": True, "type": "number",
            "label": "Default Venue Capacity",
            "description": "Default seating capacity used when a venue has no capacity specified"
        }
    },
    "gap_fill": {
        "mini_project_max_periods": {
            "value": 4, "enabled": True, "type": "number",
            "label": "Mini Project Max Periods/Week",
            "description": "Maximum periods per week allocated for mini projects during gap filling"
        },
        "core_extra_max_per_week": {
            "value": 3, "enabled": True, "type": "number",
            "label": "Core Extra Sessions/Week",
            "description": "Maximum extra gap-fill sessions per core course per week"
        },
        "core_extra_max_per_day": {
            "value": 2, "enabled": True, "type": "number",
            "label": "Core Extra Sessions/Day",
            "description": "Maximum extra gap-fill sessions per core course per day"
        },
        "open_elective_periods": {
            "value": 3, "enabled": True, "type": "number",
            "label": "Open Elective Periods (Sem 5)",
            "description": "Number of periods injected for open electives in Semester 5"
        }
    },
    "batch_rotation": {
        "enabled": {
            "value": True, "enabled": True, "type": "boolean",
            "label": "Enable Lab Batch Rotation",
            "description": "Automatically merge and rotate lab batches when there aren't enough unique lab faculty"
        },
        "venue_aware_rotation": {
            "value": True, "enabled": True, "type": "boolean",
            "label": "Venue-Aware Batch Rotation",
            "description": "Also trigger batch rotation when mapped lab venues are fewer than required sections"
        },
        "max_merged_entries": {
            "value": 15, "enabled": True, "type": "number",
            "label": "Max Merged Entries Per Slot",
            "description": "Maximum number of entries allowed in a single merged lab slot"
        }
    },
    "elective_handling": {
        "pair_same_category": {
            "value": True, "enabled": True, "type": "boolean",
            "label": "Pair Same-Category Electives",
            "description": "Group electives with the same course_category into the same time slot"
        },
        "skip_no_faculty_lang": {
            "value": True, "enabled": True, "type": "boolean",
            "label": "Skip Language Electives Without Faculty",
            "description": "Automatically skip language elective courses that have no faculty mapped"
        }
    },
    "honours_minor": {
        "slot_restriction": {
            "value": 8, "enabled": True, "type": "number",
            "label": "Honours/Minor Period",
            "description": "Period number exclusively used for honours and minor courses"
        }
    }
}

def merge_configs(default_c, saved_c):
    import copy
    result = copy.deepcopy(default_c)
    if not isinstance(saved_c, dict):
        return result
    for category, items in saved_c.items():
        if category in result and isinstance(items, dict):
            for key, val in items.items():
                if key in result[category] and isinstance(val, dict):
                    # Only merge user-editable state, keep UI schema from defaults
                    for state_key in ['value', 'enabled', 'overloaded_value']:
                        if state_key in val:
                            result[category][key][state_key] = val[state_key]
                else:
                    result[category][key] = val
        else:
            result[category] = items
    return result

@app.get("/api/config")
def get_scheduler_config(db: Session = Depends(get_db)):
    row = db.query(models.SchedulerConfig).filter_by(id=1).first()
    if not row:
        # Create defaults
        row = models.SchedulerConfig(id=1, config_json=json.dumps(DEFAULT_CONFIG))
        db.add(row)
        db.commit()
        db.refresh(row)
        return json.loads(row.config_json)
    
    saved_config = json.loads(row.config_json) if row.config_json else {}
    merged_config = merge_configs(DEFAULT_CONFIG, saved_config)
    return merged_config


@app.put("/api/config")
def update_scheduler_config(config: dict, db: Session = Depends(get_db)):
    row = db.query(models.SchedulerConfig).filter_by(id=1).first()
    if not row:
        row = models.SchedulerConfig(id=1, config_json=json.dumps(config))
        db.add(row)
    else:
        row.config_json = json.dumps(config)
    db.commit()
    return {"status": "saved"}


@app.post("/api/config/reset")
def reset_scheduler_config(db: Session = Depends(get_db)):
    row = db.query(models.SchedulerConfig).filter_by(id=1).first()
    if row:
        row.config_json = json.dumps(DEFAULT_CONFIG)
    else:
        row = models.SchedulerConfig(id=1, config_json=json.dumps(DEFAULT_CONFIG))
        db.add(row)
    db.commit()
    return json.loads(row.config_json)


# ============================================
# COMMON COURSES
# ============================================

@app.get("/common-courses")
def get_common_courses(db: Session = Depends(get_db)):
    """Return all common course groups with venue info and per-dept student counts."""
    rows = db.query(models.CommonCourseMap).all()
    # Group by course_code + semester
    groups: dict = {}
    for row in rows:
        key = (row.course_code, row.semester)
        if key not in groups:
            groups[key] = {
                "course_code": row.course_code,
                "semester": row.semester,
                "departments": [],
                "venue_name": row.venue_name,
                "venue_type": row.venue_type or 'BOTH',
                "dept_student_counts": []
            }
        # Per-dept student count
        sem_count = db.query(models.DepartmentSemesterCount).filter_by(
            department_code=row.department_code, semester=row.semester
        ).first()
        count = sem_count.student_count if sem_count and sem_count.student_count > 0 else 0
        
        groups[key]["departments"].append(row.department_code)
        groups[key]["dept_student_counts"].append({
            "dept": row.department_code,
            "count": count
        })
        # Use venue from any row (all should be same)
        if row.venue_name and not groups[key]["venue_name"]:
            groups[key]["venue_name"] = row.venue_name
            groups[key]["venue_type"] = row.venue_type or 'BOTH'
    return list(groups.values())


class CommonCourseRequest(BaseModel):
    course_code: str
    semester: int
    department_codes: List[str]

@app.post("/common-courses")
def save_common_course(req: CommonCourseRequest, db: Session = Depends(get_db)):
    """Create or replace the department list for a common course group.
    Preserves the existing global venue if one was already set."""
    # Check if there's an existing venue set
    existing_venue = None
    existing_type = 'BOTH'
    old_rows = db.query(models.CommonCourseMap).filter_by(
        course_code=req.course_code, semester=req.semester
    ).all()
    for r in old_rows:
        if r.venue_name:
            existing_venue = r.venue_name
            existing_type = r.venue_type or 'BOTH'
            break

    # Delete existing entries for this course+semester
    db.query(models.CommonCourseMap).filter_by(
        course_code=req.course_code, semester=req.semester
    ).delete()
    # Insert new entries, carrying forward the global venue
    for dept in req.department_codes:
        db.add(models.CommonCourseMap(
            course_code=req.course_code,
            semester=req.semester,
            department_code=dept,
            venue_name=existing_venue,
            venue_type=existing_type
        ))
    db.commit()
    return {"status": "saved", "course_code": req.course_code, "semester": req.semester}


class CommonCourseVenueRequest(BaseModel):
    course_code: str
    semester: int
    venue_name: str
    venue_type: str = 'BOTH'  # THEORY / LAB / BOTH

@app.post("/common-courses/venue")
def set_common_course_venue(req: CommonCourseVenueRequest, db: Session = Depends(get_db)):
    """Globally set the venue for a common course — applies to ALL departments."""
    rows = db.query(models.CommonCourseMap).filter_by(
        course_code=req.course_code, semester=req.semester
    ).all()
    if not rows:
        raise HTTPException(status_code=404, detail=f"No common course found for {req.course_code} semester {req.semester}")
    
    vtype = (req.venue_type or 'BOTH').upper()
    # Update all rows atomically
    db.query(models.CommonCourseMap).filter_by(
        course_code=req.course_code, semester=req.semester
    ).update({"venue_name": req.venue_name, "venue_type": vtype})
    
    # Also sync existing TimetableEntry rows for this common course
    db.query(models.TimetableEntry).filter_by(
        course_code=req.course_code, semester=req.semester
    ).update({"venue_name": req.venue_name})
    
    db.commit()
    
    dept_codes = [r.department_code for r in rows]
    return {
        "status": "saved",
        "course_code": req.course_code,
        "semester": req.semester,
        "venue_name": req.venue_name,
        "venue_type": vtype,
        "synced_departments": dept_codes
    }


@app.delete("/common-courses/venue/{course_code:path}/{semester}")
def clear_common_course_venue(course_code: str, semester: int, db: Session = Depends(get_db)):
    """Remove the global venue assignment from a common course."""
    affected = db.query(models.CommonCourseMap).filter_by(
        course_code=course_code, semester=semester
    ).update({"venue_name": None, "venue_type": "BOTH"})
    db.commit()
    return {"status": "cleared", "rows": affected}


@app.get("/common-courses/student-distribution/{course_code:path}/{semester}")
def get_common_course_students(course_code: str, semester: int, db: Session = Depends(get_db)):
    """Get per-department student counts for a common course."""
    rows = db.query(models.CommonCourseMap).filter_by(
        course_code=course_code, semester=semester
    ).all()
    if not rows:
        raise HTTPException(status_code=404, detail="Common course not found")
    
    breakdown = []
    total = 0
    for row in rows:
        sem_count = db.query(models.DepartmentSemesterCount).filter_by(
            department_code=row.department_code, semester=semester
        ).first()
        count = sem_count.student_count if sem_count and sem_count.student_count > 0 else 0
        breakdown.append({"dept": row.department_code, "count": count})
        total += count
    
    return {"course_code": course_code, "semester": semester, "total": total, "departments": breakdown}


@app.delete("/common-courses/{course_code:path}/{semester}")
def delete_common_course(course_code: str, semester: int, db: Session = Depends(get_db)):
    """Remove all department mappings for a common course."""
    deleted = db.query(models.CommonCourseMap).filter_by(
        course_code=course_code, semester=semester
    ).delete()
    db.commit()
    return {"status": "deleted", "rows": deleted}


# ============================================
# USER-DEFINED CONSTRAINTS
# ============================================
import uuid as uuid_lib
from datetime import datetime

def _constraint_to_response(row: models.UserConstraint) -> dict:
    """Convert a UserConstraint DB row to a response dict."""
    return {
        "id": row.id,
        "uuid": row.uuid,
        "name": row.name,
        "description": row.description,
        "enabled": row.enabled,
        "priority": row.priority,
        "soft_weight": row.soft_weight,
        "constraint_type": row.constraint_type,
        "scope": json.loads(row.scope_json),
        "target": json.loads(row.target_json),
        "rules": json.loads(row.rules_json),
        "created_at": row.created_at,
        "updated_at": row.updated_at,
        "order_index": row.order_index,
    }


@app.get("/api/user-constraints")
def list_user_constraints(
    dept: Optional[str] = None,
    sem: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """List all user constraints, optionally filtered by scope (department/semester)."""
    rows = db.query(models.UserConstraint).order_by(
        models.UserConstraint.priority.desc(),
        models.UserConstraint.order_index
    ).all()

    results = []
    for row in rows:
        resp = _constraint_to_response(row)
        # Optional filtering by scope
        if dept or sem:
            scope = resp["scope"]
            scope_depts = scope.get("departments", ["*"])
            scope_sems = scope.get("semesters", ["*"])
            if dept and scope_depts != ["*"] and dept not in scope_depts:
                continue
            if sem is not None and scope_sems != ["*"] and sem not in scope_sems:
                continue
        results.append(resp)

    return results


@app.post("/api/user-constraints")
def create_user_constraint(
    data: schemas.UserConstraintCreate,
    db: Session = Depends(get_db)
):
    """Create a new user-defined constraint."""
    now = datetime.utcnow().isoformat()
    max_order = db.query(models.UserConstraint).count()

    row = models.UserConstraint(
        uuid=str(uuid_lib.uuid4()),
        name=data.name,
        description=data.description,
        enabled=data.enabled,
        priority=data.priority,
        soft_weight=data.soft_weight,
        constraint_type=data.constraint_type,
        scope_json=json.dumps(data.scope),
        target_json=json.dumps(data.target),
        rules_json=json.dumps(data.rules),
        created_at=now,
        updated_at=now,
        order_index=max_order,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _constraint_to_response(row)


@app.put("/api/user-constraints/{constraint_uuid}")
def update_user_constraint(
    constraint_uuid: str,
    data: schemas.UserConstraintUpdate,
    db: Session = Depends(get_db)
):
    """Update an existing user-defined constraint."""
    row = db.query(models.UserConstraint).filter_by(uuid=constraint_uuid).first()
    if not row:
        raise HTTPException(status_code=404, detail="Constraint not found")

    now = datetime.utcnow().isoformat()
    if data.name is not None:
        row.name = data.name
    if data.description is not None:
        row.description = data.description
    if data.enabled is not None:
        row.enabled = data.enabled
    if data.priority is not None:
        row.priority = data.priority
    if data.soft_weight is not None:
        row.soft_weight = data.soft_weight
    if data.constraint_type is not None:
        row.constraint_type = data.constraint_type
    if data.scope is not None:
        row.scope_json = json.dumps(data.scope)
    if data.target is not None:
        row.target_json = json.dumps(data.target)
    if data.rules is not None:
        row.rules_json = json.dumps(data.rules)
    row.updated_at = now

    db.commit()
    db.refresh(row)
    return _constraint_to_response(row)


@app.delete("/api/user-constraints/{constraint_uuid}")
def delete_user_constraint(
    constraint_uuid: str,
    db: Session = Depends(get_db)
):
    """Delete a user-defined constraint."""
    deleted = db.query(models.UserConstraint).filter_by(uuid=constraint_uuid).delete()
    if not deleted:
        raise HTTPException(status_code=404, detail="Constraint not found")
    db.commit()
    return {"status": "deleted", "uuid": constraint_uuid}


@app.patch("/api/user-constraints/{constraint_uuid}/toggle")
def toggle_user_constraint(
    constraint_uuid: str,
    db: Session = Depends(get_db)
):
    """Toggle the enabled status of a constraint."""
    row = db.query(models.UserConstraint).filter_by(uuid=constraint_uuid).first()
    if not row:
        raise HTTPException(status_code=404, detail="Constraint not found")
    row.enabled = not row.enabled
    row.updated_at = datetime.utcnow().isoformat()
    db.commit()
    db.refresh(row)
    return _constraint_to_response(row)


@app.post("/api/user-constraints/reorder")
def reorder_user_constraints(
    data: schemas.UserConstraintReorderRequest,
    db: Session = Depends(get_db)
):
    """Update the display/processing order of constraints."""
    for idx, c_uuid in enumerate(data.order):
        row = db.query(models.UserConstraint).filter_by(uuid=c_uuid).first()
        if row:
            row.order_index = idx
    db.commit()
    return {"status": "reordered"}


@app.post("/api/user-constraints/validate")
def validate_user_constraint(
    data: schemas.UserConstraintCreate,
    db: Session = Depends(get_db)
):
    """Validate a constraint definition: check for obvious issues without running the solver."""
    warnings = []

    # Basic validation
    valid_types = ['COURSE_INJECTION', 'SLOT_BLOCKING', 'FACULTY_RULE',
                   'SPACING_RULE', 'DISTRIBUTION_RULE', 'CAPACITY_RULE', 'CUSTOM_PLACEMENT']
    if data.constraint_type not in valid_types:
        warnings.append(f"Unknown constraint type: {data.constraint_type}")

    valid_priorities = ['HARD', 'SOFT', 'PREFERENCE']
    if data.priority not in valid_priorities:
        warnings.append(f"Unknown priority: {data.priority}")

    # Scope validation
    scope = data.scope or {}
    scope_depts = scope.get("departments", ["*"])
    if scope_depts != ["*"]:
        for dept in scope_depts:
            if not db.query(models.DepartmentMaster).filter_by(department_code=dept).first():
                warnings.append(f"Department '{dept}' not found")

    # Target validation
    target = data.target or {}
    if target.get("type") == "COURSE" and target.get("course_code"):
        if not target.get("create_if_missing"):
            exists = db.query(models.CourseMaster).filter_by(
                course_code=target["course_code"]
            ).first()
            if not exists:
                warnings.append(f"Course '{target['course_code']}' not found. Enable 'create_if_missing' to auto-create.")

    if target.get("type") == "FACULTY" and target.get("faculty_id"):
        exists = db.query(models.FacultyMaster).filter_by(
            faculty_id=target["faculty_id"]
        ).first()
        if not exists:
            warnings.append(f"Faculty '{target['faculty_id']}' not found")

    # Rules validation
    rules = data.rules or {}
    sessions = rules.get("sessions_per_week", {})
    if sessions:
        min_s = sessions.get("min", 0)
        max_s = sessions.get("max", 99)
        exact = sessions.get("exact")
        if exact and exact > 7 * 8:  # 7 days * 8 periods max
            warnings.append(f"Requested {exact} sessions/week exceeds maximum possible slots")
        if min_s > max_s:
            warnings.append(f"sessions_per_week min ({min_s}) > max ({max_s})")

    # Visual slots validation
    visual_slots = rules.get("visual_slots", [])
    if visual_slots:
        for vs in visual_slots:
            slot = db.query(models.SlotMaster).filter_by(
                day_of_week=vs.get("day"),
                period_number=vs.get("period")
            ).first()
            if not slot:
                warnings.append(f"Slot {vs.get('day')} P{vs.get('period')} does not exist")

    return {
        "valid": len(warnings) == 0,
        "warnings": warnings
    }

# ============================================
# EXPORT
# ============================================

@app.get("/export/timetable/excel")
def export_timetable_excel(
    department_code: str = None, 
    semester: int = None, 
    db: Session = Depends(get_db)
):
    """Generates a highly-formatted Excel (.xlsx) version of the Timetable."""
    # Fetch slots
    slots = db.query(models.SlotMaster).order_by(models.SlotMaster.day_of_week, models.SlotMaster.period_number).all()
    if not slots:
        raise HTTPException(status_code=400, detail="No slots configured")
        
    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    active_days = [d for d in days_order if any(s.day_of_week == d for s in slots)]
    if not active_days: active_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    
    max_period = max([s.period_number for s in slots], default=8)
    periods = list(range(1, max_period + 1))
    
    # Fetch timetable entries
    query = db.query(models.TimetableEntry)
    if department_code: query = query.filter(models.TimetableEntry.department_code == department_code)
    if semester: query = query.filter(models.TimetableEntry.semester == semester)
    entries = query.all()
    
    # Group entries by (day, period)
    timetable_dict = {}
    for e in entries:
        key = (e.day_of_week, e.period_number)
        if key not in timetable_dict:
            timetable_dict[key] = []
        timetable_dict[key].append(e)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Timetable"
    
    # Styles matching PDF layout exactly
    thin = Side(border_style="thin", color="D1D5DB") # Gray-300
    thick = Side(border_style="medium", color="1F2937") # Gray-800
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    fill_header = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # bg-gray-50
    fill_day = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # bg-gray-50
    fill_lab = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # bg-amber-50 equivalent
    fill_class = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_empty = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # bg-gray-100
    
    font_bold = Font(bold=True, color="1F2937", size=11)
    font_title = Font(bold=True, size=16, color="111827")
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    try:
        from openpyxl.drawing.image import Image
    except ImportError:
        Image = None
    import os

    # Fetch Academic Year (global - stored at semester=0)
    academic_year = "___________"
    config = db.query(models.SemesterConfig).filter_by(semester=0).first()
    if config and config.academic_year:
        academic_year = config.academic_year
    
    # Fetch active breaks
    all_breaks = db.query(models.BreakConfigMaster).all()
    active_breaks = []
    for b in all_breaks:
        sem_ids = json.loads(b.semester_ids) if b.semester_ids else []
        if not sem_ids or (semester and semester in sem_ids):
            active_breaks.append(b)

    fill_break = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    header_cols_total = len(periods) + 1

    # --- ROW 1-5: INSTITUTION HEADER ---
    logo_path = os.path.join(os.path.dirname(__file__), "bitsathy-logo.png")
    if Image and os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            img.width = 100
            img.height = 100
            ws.add_image(img, "A1")
        except Exception as e:
            print(f"Warning: Could not inject logo: {e}")

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=header_cols_total)
    c1 = ws.cell(row=1, column=2, value="BANNARI AMMAN INSTITUTE OF TECHNOLOGY")
    c1.font = Font(bold=True, size=18, color="1E3A8A")
    c1.alignment = align_center

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=header_cols_total)
    c2 = ws.cell(row=2, column=2, value="(An Autonomous Institution Affiliated to Anna University, Chennai)")
    c2.font = Font(bold=True, size=11, color="111827")
    c2.alignment = align_center

    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=header_cols_total)
    c3 = ws.cell(row=3, column=2, value="Sathyamangalam - 638401")
    c3.font = Font(bold=True, size=11, color="111827")
    c3.alignment = align_center

    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=header_cols_total)
    c4 = ws.cell(row=4, column=2, value="CLASS TIMETABLE")
    c4.font = Font(bold=True, size=14, color="111827")
    c4.alignment = align_center

    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=header_cols_total)
    c5 = ws.cell(row=5, column=2, value=f"Academic Year {academic_year}")
    c5.font = Font(bold=True, size=12, color="111827")
    c5.alignment = align_center

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20

    # --- ROW 6-7: BREAK DISPLAY HEADER ---
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=header_cols_total)
    b_head = ws.cell(row=6, column=1, value="SCHEDULED BREAKS")
    b_head.font = Font(bold=True, size=10, color="1F2937")
    b_head.fill = fill_day
    b_head.alignment = align_center
    b_head.border = border_all

    break_text = "  |  ".join([f"{b.break_type}: {b.start_time}-{b.end_time}" for b in active_breaks]) if active_breaks else "No breaks scheduled"
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=header_cols_total)
    b_val = ws.cell(row=7, column=1, value=break_text)
    b_val.font = Font(bold=True, size=10, color="4B5563")
    b_val.fill = fill_break
    b_val.alignment = align_center
    b_val.border = border_all

    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 25

    # --- ROW 8: DEPARTMENT AND SEMESTER META ROW ---
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=header_cols_total)
    meta_text = f"DEPARTMENT OF {department_code or 'ALL'} | SEMESTER {semester or 'ALL'}"
    m_cell = ws.cell(row=8, column=1, value=meta_text)
    m_cell.font = Font(bold=True, size=12, color="000000")
    m_cell.alignment = align_center
    ws.row_dimensions[8].height = 25
    
    # Space between title and table
    header_row = 10
    
    # Headers Setup
    d_cell = ws.cell(row=header_row, column=1, value="DAY / PERIOD")
    d_cell.font = Font(bold=True, color="1F2937", size=10)
    d_cell.alignment = align_center
    d_cell.fill = fill_day
    d_cell.border = border_all
    ws.column_dimensions['A'].width = 16
    
    for i, p in enumerate(periods):
        col = i + 2
        slot = next((s for s in slots if s.period_number == p and s.day_of_week == 'Monday'), None)
        time_str = f"{slot.start_time} - {slot.end_time}" if slot else ""
        header_text = f"Period {p}\n({time_str})"
        c = ws.cell(row=header_row, column=col, value=header_text)
        c.font = Font(bold=True, color="374151", size=10)
        c.alignment = align_center
        c.fill = fill_header
        c.border = border_all
        ws.column_dimensions[get_column_letter(col)].width = 24
        
    ws.row_dimensions[header_row].height = 35
    
    # Body Setup
    current_row = header_row + 1
    for day in active_days:
        dc = ws.cell(row=current_row, column=1, value=day[:3].upper()) # MON, TUE etc.
        dc.font = Font(bold=True, size=12, color="1F2937")
        dc.alignment = align_center
        dc.fill = fill_day
        dc.border = border_all
        
        ws.row_dimensions[current_row].height = 85
        
        skip_next = False
        for i, p in enumerate(periods):
            col = i + 2
            if skip_next:
                skip_next = False
                continue
                
            cell_entries = timetable_dict.get((day, p), [])
            # Check for merged LAB span
            is_lab = any(e.session_type.upper() == 'LAB' for e in cell_entries) if cell_entries else False
            col_span = 1
            if is_lab and i < len(periods) - 1:
                next_p = periods[i+1]
                next_entries = timetable_dict.get((day, next_p), [])
                if any(e.session_type.upper() == 'LAB' for e in next_entries):
                    cur_codes = sorted(list(set(e.course_code for e in cell_entries if e.session_type.upper() == 'LAB')))
                    nxt_codes = sorted(list(set(e.course_code for e in next_entries if e.session_type.upper() == 'LAB')))
                    if cur_codes == nxt_codes and len(cur_codes) > 0:
                        col_span = 2
                        skip_next = True
            
            # Text Construction
            if not cell_entries:
                cell_text = "Empty Slot"
                fill = fill_empty
                cell_font = Font(italic=True, color="9CA3AF", size=10)
            else:
                lines = []
                for idx, e in enumerate(cell_entries):
                    c_name = e.course_name or ''
                    fac = e.faculty_name or 'Unassigned'
                    ven = e.venue_name or ''
                    block = f"{e.course_code}\n{c_name}\n({fac})\n[{ven}]"
                    lines.append(block)
                cell_text = "\n\n---\n\n".join(lines)
                fill = fill_lab if is_lab else fill_class
                cell_font = Font(bold=False, size=10, color="111827")
                
            c = ws.cell(row=current_row, column=col, value=cell_text)
            c.alignment = align_center
            c.fill = fill
            c.border = border_all
            c.font = cell_font
            
            if col_span > 1:
                ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + col_span - 1)
                for merge_col in range(col, col + col_span):
                    mc = ws.cell(row=current_row, column=merge_col)
                    mc.border = border_all
                    
        current_row += 1

    # Apply thick borders to outline the table
    for r in range(header_row, current_row):
        ws.cell(row=r, column=1).border = Border(left=thick, top=ws.cell(row=r, column=1).border.top, bottom=ws.cell(row=r, column=1).border.bottom, right=ws.cell(row=r, column=1).border.right)
        ws.cell(row=r, column=len(periods)+1).border = Border(right=thick, top=ws.cell(row=r, column=len(periods)+1).border.top, bottom=ws.cell(row=r, column=len(periods)+1).border.bottom, left=ws.cell(row=r, column=len(periods)+1).border.left)
    for c in range(1, len(periods)+2):
        ws.cell(row=header_row, column=c).border = Border(top=thick, left=ws.cell(row=header_row, column=c).border.left, right=ws.cell(row=header_row, column=c).border.right, bottom=ws.cell(row=header_row, column=c).border.bottom)
        ws.cell(row=current_row-1, column=c).border = Border(bottom=thick, left=ws.cell(row=current_row-1, column=c).border.left, right=ws.cell(row=current_row-1, column=c).border.right, top=ws.cell(row=current_row-1, column=c).border.top)
    
    ws.cell(row=header_row, column=1).border = Border(top=thick, left=thick, bottom=ws.cell(row=header_row, column=1).border.bottom, right=ws.cell(row=header_row, column=1).border.right)
    ws.cell(row=header_row, column=len(periods)+1).border = Border(top=thick, right=thick, bottom=ws.cell(row=header_row, column=len(periods)+1).border.bottom, left=ws.cell(row=header_row, column=len(periods)+1).border.left)
    ws.cell(row=current_row-1, column=1).border = Border(bottom=thick, left=thick, top=ws.cell(row=current_row-1, column=1).border.top, right=ws.cell(row=current_row-1, column=1).border.right)
    ws.cell(row=current_row-1, column=len(periods)+1).border = Border(bottom=thick, right=thick, top=ws.cell(row=current_row-1, column=len(periods)+1).border.top, left=ws.cell(row=current_row-1, column=len(periods)+1).border.left)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="timetable_{department_code or "all"}_{semester or "all"}.xlsx"'
    }
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
