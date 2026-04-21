"""
Legacy routes — ported from old/backend/main.py
All working endpoints consolidated into a single APIRouter.
Mounted at /api prefix in main.py so all paths are /api/...
"""

from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, Query
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
import json
import math
import io
import re
import os
import copy
import uuid as uuid_lib
from datetime import datetime

from utils.database import get_db
import models
import schemas

router = APIRouter()


# ============================================
# HEALTH
# ============================================

@router.get("/health")
def health_check():
    return {"status": "ok"}


# ============================================
# SYNC CMS
# ============================================

SYNC_STATE = {
    "is_running": False,
    "status": "idle",
    "error": None
}

import asyncio

async def _run_sync_job():
    global SYNC_STATE
    SYNC_STATE["is_running"] = True
    SYNC_STATE["status"] = "syncing"
    SYNC_STATE["error"] = None
    SYNC_STATE["result"] = None
    try:
        from services.sync_cms_to_local import sync_databases
        result = await asyncio.to_thread(sync_databases)
        SYNC_STATE["status"] = "complete"
        SYNC_STATE["result"] = result
    except Exception as e:
        SYNC_STATE["status"] = "error"
        SYNC_STATE["error"] = str(e)
    finally:
        SYNC_STATE["is_running"] = False

@router.post("/sync-cms")
async def run_sync_cms(background_tasks: BackgroundTasks):
    global SYNC_STATE
    if SYNC_STATE["is_running"]:
        return {"status": "started", "message": "A sync is already in progress."}
    # FastAPI background tasks execute in threadpool, ensuring event loop is free
    background_tasks.add_task(_run_sync_job)
    return {"status": "started", "message": "CMS Data sync started in the background."}

@router.get("/sync-cms/status")
def get_sync_status():
    global SYNC_STATE
    return SYNC_STATE


# ============================================
# SEMESTER CONFIG
# ============================================

@router.get("/semester-config", response_model=List[schemas.SemesterConfigResponse])
def get_semester_configs(db: Session = Depends(get_db)):
    return db.query(models.SemesterConfig).all()

@router.post("/semester-config/{semester}", response_model=schemas.SemesterConfigResponse)
def update_semester_config(semester: int, req: schemas.SemesterConfigUpdate, db: Session = Depends(get_db)):
    config = db.query(models.SemesterConfig).filter_by(semester=semester).first()
    if not config:
        config = models.SemesterConfig(semester=semester, academic_year=req.academic_year)
        db.add(config)
    else:
        config.academic_year = req.academic_year
    db.commit()
    db.refresh(config)
    return config


# ============================================
# DEPARTMENTS
# ============================================

@router.get("/departments", response_model=List[schemas.Department])
def get_departments(db: Session = Depends(get_db)):
    return db.query(models.DepartmentMaster).all()

@router.post("/departments")
def create_department(req: schemas.DepartmentCreate, db: Session = Depends(get_db)):
    existing = db.query(models.DepartmentMaster).filter_by(department_code=req.department_code).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Department {req.department_code} already exists")
    dept = models.DepartmentMaster(
        department_code=req.department_code,
        student_count=req.student_count,
        pair_add_course_miniproject=req.pair_add_course_miniproject
    )
    db.add(dept)
    db.commit()
    return {"status": "success", "department_code": req.department_code}

@router.put("/departments/{code}", response_model=schemas.Department)
def update_department(code: str, req: schemas.DepartmentUpdate, db: Session = Depends(get_db)):
    dept = db.query(models.DepartmentMaster).filter_by(department_code=code).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    if req.student_count is not None:
        dept.student_count = req.student_count
    if req.pair_add_course_miniproject is not None:
        dept.pair_add_course_miniproject = req.pair_add_course_miniproject
    db.commit()
    db.refresh(dept)
    return dept

@router.delete("/departments/{code}")
def delete_department(code: str, db: Session = Depends(get_db)):
    dept = db.query(models.DepartmentMaster).filter_by(department_code=code).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    fac_count = db.query(models.FacultyMaster).filter_by(department_code=code).count()
    course_count = db.query(models.CourseMaster).filter_by(department_code=code).count()
    if fac_count > 0 or course_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {fac_count} faculty and {course_count} courses still linked.")
    db.delete(dept)
    db.commit()
    return {"status": "deleted"}

@router.get("/departments/{code}/capacities", response_model=List[schemas.DepartmentSemesterCountResponse])
def get_department_capacities(code: str, db: Session = Depends(get_db)):
    dept = db.query(models.DepartmentMaster).filter_by(department_code=code).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    return db.query(models.DepartmentSemesterCount).filter_by(department_code=code).order_by(models.DepartmentSemesterCount.semester).all()

@router.post("/departments/{code}/capacities", response_model=schemas.DepartmentSemesterCountResponse)
def upsert_department_capacity(code: str, semester: int = Query(...), req: schemas.DepartmentSemesterCountUpdate = ..., db: Session = Depends(get_db)):
    dept = db.query(models.DepartmentMaster).filter_by(department_code=code).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    record = db.query(models.DepartmentSemesterCount).filter_by(department_code=code, semester=semester).first()
    if record:
        record.student_count = req.student_count
    else:
        record = models.DepartmentSemesterCount(department_code=code, semester=semester, student_count=req.student_count)
        db.add(record)
    db.commit()
    db.refresh(record)
    return record


# ============================================
# SEMESTERS
# ============================================

@router.get("/semesters")
def get_semesters():
    return [{"semester_number": i} for i in range(1, 9)]


# ============================================
# FACULTY
# ============================================

@router.get("/faculty", response_model=List[schemas.Faculty])
def get_faculty(department_code: str = None, db: Session = Depends(get_db)):
    query = db.query(models.FacultyMaster)
    if department_code:
        query = query.filter(models.FacultyMaster.department_code == department_code)
    return query.all()

@router.post("/faculty")
def create_faculty(req: schemas.FacultyCreate, db: Session = Depends(get_db)):
    existing = db.query(models.FacultyMaster).filter_by(faculty_id=req.faculty_id).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Faculty {req.faculty_id} already exists")
    dept = db.query(models.DepartmentMaster).filter_by(department_code=req.department_code).first()
    if not dept:
        raise HTTPException(status_code=400, detail=f"Department {req.department_code} does not exist")
    fac = models.FacultyMaster(
        faculty_id=req.faculty_id,
        faculty_name=req.faculty_name,
        faculty_email=req.faculty_email,
        department_code=req.department_code,
        status=req.status
    )
    db.add(fac)
    db.commit()
    return {"status": "success", "faculty_id": req.faculty_id}

@router.delete("/faculty/{fid}")
def delete_faculty(fid: str, db: Session = Depends(get_db)):
    fac = db.query(models.FacultyMaster).filter_by(faculty_id=fid).first()
    if not fac:
        raise HTTPException(status_code=404, detail="Faculty not found")
    db.query(models.CourseFacultyMap).filter_by(faculty_id=fid).delete()
    db.delete(fac)
    db.commit()
    return {"status": "deleted"}


# ============================================
# COURSES
# ============================================

@router.get("/courses", response_model=List[schemas.Course])
def get_courses(department_code: str = None, semester: int = None, db: Session = Depends(get_db)):
    query = db.query(models.CourseMaster)
    if department_code:
        query = query.filter(models.CourseMaster.department_code == department_code)
    if semester:
        query = query.filter(models.CourseMaster.semester == semester)
    return query.all()

@router.post("/courses")
def create_course(req: schemas.CourseCreate, db: Session = Depends(get_db)):
    target_depts = [req.department_code]
    if getattr(req, 'common_departments', None):
        for d in req.common_departments:
            if d not in target_depts:
                target_depts.append(d)
    for d_code in target_depts:
        dept = db.query(models.DepartmentMaster).filter_by(department_code=d_code).first()
        if not dept:
            raise HTTPException(status_code=400, detail=f"Department {d_code} does not exist")
    is_honours = req.is_honours or ('H' in req.course_code[-3:].upper())
    created_count = 0
    for d_code in target_depts:
        existing = db.query(models.CourseMaster).filter_by(course_code=req.course_code, department_code=d_code).first()
        if existing:
            if d_code == req.department_code:
                raise HTTPException(status_code=400, detail=f"Course {req.course_code} already exists in {d_code}")
            continue
        course = models.CourseMaster(
            course_code=req.course_code, course_name=req.course_name, department_code=d_code,
            semester=req.semester, lecture_hours=req.lecture_hours, tutorial_hours=req.tutorial_hours,
            practical_hours=req.practical_hours, credits=req.credits, weekly_sessions=req.weekly_sessions,
            is_lab=req.is_lab, is_honours=is_honours, is_minor=req.is_minor, is_elective=req.is_elective,
            is_open_elective=req.is_open_elective, is_add_course=req.is_add_course
        )
        db.add(course)
        created_count += 1
    if len(target_depts) > 1:
        db.query(models.CommonCourseMap).filter_by(course_code=req.course_code, semester=req.semester).delete()
        for d_code in target_depts:
            db.add(models.CommonCourseMap(course_code=req.course_code, semester=req.semester, department_code=d_code))
    db.commit()
    return {"status": "success", "course_code": req.course_code, "created_records": created_count}

@router.put("/courses/{code:path}")
def update_course(code: str, req: schemas.CourseUpdate, db: Session = Depends(get_db)):
    existing_courses = db.query(models.CourseMaster).filter_by(course_code=code).all()
    if not existing_courses:
        raise HTTPException(status_code=404, detail="Course not found")
    base_course = existing_courses[0]
    new_code = req.course_code if req.course_code is not None else code
    if new_code != code:
        conflict = db.query(models.CourseMaster).filter_by(course_code=new_code).first()
        if conflict:
            raise HTTPException(status_code=400, detail=f"Course {new_code} already exists")
        for existing in existing_courses:
            db.add(models.CourseMaster(
                course_code=new_code, department_code=existing.department_code, semester=existing.semester,
                course_name=existing.course_name, course_category=existing.course_category,
                delivery_type=existing.delivery_type, lecture_hours=existing.lecture_hours,
                tutorial_hours=existing.tutorial_hours, practical_hours=existing.practical_hours,
                credits=existing.credits, weekly_sessions=existing.weekly_sessions, is_lab=existing.is_lab,
                is_elective=existing.is_elective, is_open_elective=existing.is_open_elective,
                is_honours=existing.is_honours, is_minor=existing.is_minor, is_add_course=existing.is_add_course,
                enrolled_students=existing.enrolled_students
            ))
        db.flush()
        db.query(models.CourseFacultyMap).filter_by(course_code=code).update({"course_code": new_code})
        db.query(models.CourseVenueMap).filter_by(course_code=code).update({"course_code": new_code})
        db.query(models.CommonCourseMap).filter_by(course_code=code).update({"course_code": new_code})
        db.query(models.CourseRegistration).filter_by(course_code=code).update({"course_code": new_code})
        db.query(models.TimetableEntry).filter_by(course_code=code).update({"course_code": new_code})
        for existing in existing_courses:
            db.delete(existing)
        db.flush()
        existing_courses = db.query(models.CourseMaster).filter_by(course_code=new_code).all()
        base_course = existing_courses[0]
        code = new_code

    primary_dept = req.department_code if req.department_code is not None else base_course.department_code
    semester = req.semester if req.semester is not None else base_course.semester
    target_depts = [primary_dept]
    if hasattr(req, 'common_departments') and req.common_departments is not None:
        for d in req.common_departments:
            if d not in target_depts:
                target_depts.append(d)
    else:
        common_maps = db.query(models.CommonCourseMap).filter_by(course_code=code).all()
        for cm in common_maps:
            if cm.department_code not in target_depts:
                target_depts.append(cm.department_code)
    for d_code in target_depts:
        if not db.query(models.DepartmentMaster).filter_by(department_code=d_code).first():
            raise HTTPException(status_code=400, detail=f"Department {d_code} does not exist")
    for existing in existing_courses:
        if existing.department_code not in target_depts:
            db.delete(existing)
    for d_code in target_depts:
        course = db.query(models.CourseMaster).filter_by(course_code=code, department_code=d_code).first()
        if not course:
            course = models.CourseMaster(course_code=code, department_code=d_code)
            db.add(course)
        if req.course_name is not None: course.course_name = req.course_name
        if req.semester is not None: course.semester = req.semester
        if req.lecture_hours is not None: course.lecture_hours = req.lecture_hours
        if req.tutorial_hours is not None: course.tutorial_hours = req.tutorial_hours
        if req.practical_hours is not None: course.practical_hours = req.practical_hours
        if req.credits is not None: course.credits = req.credits
        if req.weekly_sessions is not None: course.weekly_sessions = req.weekly_sessions
        if req.is_lab is not None: course.is_lab = req.is_lab
        if req.is_honours is not None: course.is_honours = req.is_honours
        if req.is_minor is not None: course.is_minor = req.is_minor
        if req.is_elective is not None: course.is_elective = req.is_elective
        if req.is_open_elective is not None: course.is_open_elective = req.is_open_elective
        if req.is_add_course is not None: course.is_add_course = req.is_add_course
    db.query(models.CommonCourseMap).filter_by(course_code=code).delete()
    if len(target_depts) > 1:
        for d_code in target_depts:
            db.add(models.CommonCourseMap(course_code=code, semester=semester, department_code=d_code))
    db.commit()
    return {"status": "success", "course_code": code}

@router.delete("/courses/{code:path}")
def delete_course(code: str, db: Session = Depends(get_db)):
    courses = db.query(models.CourseMaster).filter_by(course_code=code).all()
    if not courses:
        raise HTTPException(status_code=404, detail="Course not found")
    db.query(models.CourseFacultyMap).filter_by(course_code=code).delete()
    db.query(models.CommonCourseMap).filter_by(course_code=code).delete()
    for c in courses:
        db.delete(c)
    db.commit()
    return {"status": "deleted"}


# ============================================
# COURSE-FACULTY MAPPING
# ============================================

@router.get("/course-faculty")
def get_course_faculty(department_code: str = None, db: Session = Depends(get_db)):
    query = db.query(models.CourseFacultyMap)
    if department_code:
        query = query.filter(models.CourseFacultyMap.department_code == department_code)
    mappings = query.all()
    result = []
    for m in mappings:
        fac = db.query(models.FacultyMaster).filter_by(faculty_id=m.faculty_id).first()
        course = db.query(models.CourseMaster).filter_by(course_code=m.course_code).first()
        result.append({
            "id": m.id, "course_code": m.course_code,
            "course_name": course.course_name if course else m.course_code,
            "course_dept": course.department_code if course else "?",
            "course_semester": course.semester if course else 0,
            "faculty_id": m.faculty_id,
            "faculty_name": fac.faculty_name if fac else m.faculty_id,
            "faculty_dept": fac.department_code if fac else "?",
            "for_department": m.department_code,
            "delivery_type": m.delivery_type,
        })
    return result

@router.post("/course-faculty")
def create_course_faculty(req: schemas.CourseFacultyCreate, db: Session = Depends(get_db)):
    course = db.query(models.CourseMaster).filter_by(course_code=req.course_code).first()
    if not course:
        raise HTTPException(status_code=400, detail=f"Course {req.course_code} does not exist")
    fac = db.query(models.FacultyMaster).filter_by(faculty_id=req.faculty_id).first()
    if not fac:
        raise HTTPException(status_code=400, detail=f"Faculty {req.faculty_id} does not exist")
    existing = db.query(models.CourseFacultyMap).filter_by(course_code=req.course_code, faculty_id=req.faculty_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="This mapping already exists")
    m = models.CourseFacultyMap(
        course_code=req.course_code, faculty_id=req.faculty_id,
        department_code=req.department_code, delivery_type=req.delivery_type
    )
    db.add(m)
    db.commit()
    return {"status": "success", "id": m.id}

@router.delete("/course-faculty/{mid}")
def delete_course_faculty(mid: int, db: Session = Depends(get_db)):
    m = db.query(models.CourseFacultyMap).filter_by(id=mid).first()
    if not m:
        raise HTTPException(status_code=404, detail="Mapping not found")
    db.delete(m)
    db.commit()
    return {"status": "deleted"}


# ============================================
# SLOTS
# ============================================

@router.get("/slots")
def get_slots(db: Session = Depends(get_db)):
    slots = db.query(models.SlotMaster).order_by(models.SlotMaster.day_of_week, models.SlotMaster.period_number).all()
    out = []
    for s in slots:
        try:
            parsed_ids = json.loads(s.semester_ids) if s.semester_ids else []
        except:
            parsed_ids = []
        out.append({
            "slot_id": s.slot_id, "day_of_week": s.day_of_week,
            "period_number": s.period_number, "start_time": s.start_time,
            "end_time": s.end_time, "slot_type": s.slot_type,
            "is_active": s.is_active, "semester_ids": parsed_ids
        })
    return out

@router.post("/slots")
def create_slot(req: schemas.SlotCreate, db: Session = Depends(get_db)):
    existing = db.query(models.SlotMaster).filter_by(day_of_week=req.day_of_week, period_number=req.period_number).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Slot already exists for {req.day_of_week} Period {req.period_number}")
    slot = models.SlotMaster(
        day_of_week=req.day_of_week, period_number=req.period_number,
        start_time=req.start_time, end_time=req.end_time,
        slot_type=req.slot_type, is_active=req.is_active,
        semester_ids=json.dumps(req.semester_ids) if req.semester_ids is not None else "[]"
    )
    db.add(slot)
    db.commit()
    db.refresh(slot)
    try: parsed_ids = json.loads(slot.semester_ids)
    except: parsed_ids = []
    return {
        "slot_id": slot.slot_id, "day_of_week": slot.day_of_week,
        "period_number": slot.period_number, "start_time": slot.start_time,
        "end_time": slot.end_time, "slot_type": slot.slot_type,
        "is_active": slot.is_active, "semester_ids": parsed_ids
    }

@router.put("/slots/{slot_id}")
def update_slot(slot_id: int, req: schemas.SlotUpdate, db: Session = Depends(get_db)):
    slot = db.query(models.SlotMaster).filter_by(slot_id=slot_id).first()
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    if req.start_time is not None: slot.start_time = req.start_time
    if req.end_time is not None: slot.end_time = req.end_time
    if req.slot_type is not None: slot.slot_type = req.slot_type
    if req.is_active is not None: slot.is_active = req.is_active
    if req.semester_ids is not None: slot.semester_ids = json.dumps(req.semester_ids)
    db.commit()
    db.refresh(slot)
    try: parsed_ids = json.loads(slot.semester_ids)
    except: parsed_ids = []
    return {
        "slot_id": slot.slot_id, "day_of_week": slot.day_of_week,
        "period_number": slot.period_number, "start_time": slot.start_time,
        "end_time": slot.end_time, "slot_type": slot.slot_type,
        "is_active": slot.is_active, "semester_ids": parsed_ids
    }

@router.delete("/slots/{slot_id}")
def delete_slot(slot_id: int, db: Session = Depends(get_db)):
    slot = db.query(models.SlotMaster).filter_by(slot_id=slot_id).first()
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    db.delete(slot)
    db.commit()
    return {"status": "deleted", "slot_id": slot_id}


# ============================================
# BREAKS
# ============================================

@router.get("/breaks")
def get_breaks(db: Session = Depends(get_db)):
    breaks = db.query(models.BreakConfigMaster).all()
    out = []
    for b in breaks:
        try:
            parsed_ids = json.loads(b.semester_ids) if b.semester_ids else []
        except:
            parsed_ids = []
        out.append({
            "id": b.id, "break_type": b.break_type,
            "start_time": b.start_time, "end_time": b.end_time,
            "semester_ids": parsed_ids
        })
    return out

@router.post("/breaks")
def create_break(req: schemas.BreakConfigCreate, db: Session = Depends(get_db)):
    b = models.BreakConfigMaster(
        break_type=req.break_type, start_time=req.start_time, end_time=req.end_time,
        semester_ids=json.dumps(req.semester_ids) if req.semester_ids is not None else "[]"
    )
    db.add(b)
    db.commit()
    db.refresh(b)
    try: parsed_ids = json.loads(b.semester_ids)
    except: parsed_ids = []
    return {"id": b.id, "break_type": b.break_type, "start_time": b.start_time, "end_time": b.end_time, "semester_ids": parsed_ids}

@router.put("/breaks/{break_id}")
def update_break(break_id: int, req: schemas.BreakConfigUpdate, db: Session = Depends(get_db)):
    b = db.query(models.BreakConfigMaster).filter_by(id=break_id).first()
    if not b:
        raise HTTPException(status_code=404, detail="Break config not found")
    if req.break_type is not None: b.break_type = req.break_type
    if req.start_time is not None: b.start_time = req.start_time
    if req.end_time is not None: b.end_time = req.end_time
    if req.semester_ids is not None: b.semester_ids = json.dumps(req.semester_ids)
    db.commit()
    db.refresh(b)
    try: parsed_ids = json.loads(b.semester_ids)
    except: parsed_ids = []
    return {"id": b.id, "break_type": b.break_type, "start_time": b.start_time, "end_time": b.end_time, "semester_ids": parsed_ids}

@router.delete("/breaks/{break_id}")
def delete_break(break_id: int, db: Session = Depends(get_db)):
    b = db.query(models.BreakConfigMaster).filter_by(id=break_id).first()
    if not b:
        raise HTTPException(status_code=404, detail="Break config not found")
    db.delete(b)
    db.commit()
    return {"status": "deleted", "id": break_id}


# ============================================
# VENUES
# ============================================

@router.get("/venues", response_model=List[schemas.Venue])
def get_venues(db: Session = Depends(get_db)):
    return db.query(models.VenueMaster).all()

@router.post("/venues", response_model=schemas.Venue)
def create_venue(req: schemas.VenueCreate, db: Session = Depends(get_db)):
    existing = db.query(models.VenueMaster).filter_by(venue_name=req.venue_name).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Venue '{req.venue_name}' already exists")
    venue = models.VenueMaster(venue_name=req.venue_name, block=req.block, is_lab=req.is_lab, capacity=req.capacity)
    db.add(venue)
    try:
        db.commit()
        db.refresh(venue)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    return venue

@router.put("/venues/{venue_id}", response_model=schemas.Venue)
def update_venue(venue_id: int, req: schemas.VenueUpdate, db: Session = Depends(get_db)):
    venue = db.query(models.VenueMaster).filter_by(venue_id=venue_id).first()
    if not venue:
        raise HTTPException(status_code=404, detail="Venue not found")
    if req.venue_name is not None:
        if req.venue_name != venue.venue_name:
            existing = db.query(models.VenueMaster).filter_by(venue_name=req.venue_name).first()
            if existing:
                raise HTTPException(status_code=400, detail=f"Venue '{req.venue_name}' already exists")
        venue.venue_name = req.venue_name
    if req.block is not None: venue.block = req.block
    if req.is_lab is not None: venue.is_lab = req.is_lab
    if req.capacity is not None: venue.capacity = req.capacity
    db.commit()
    db.refresh(venue)
    return venue

@router.delete("/venues/{venue_id}")
def delete_venue(venue_id: int, db: Session = Depends(get_db)):
    venue = db.query(models.VenueMaster).filter_by(venue_id=venue_id).first()
    if not venue:
        raise HTTPException(status_code=404, detail="Venue not found")
    db.delete(venue)
    db.commit()
    return {"status": "deleted", "venue_id": venue_id}


# ============================================
# DEPARTMENT-VENUE MAPPING
# ============================================

@router.get("/department-venues", response_model=List[schemas.DepartmentVenueResponse])
def get_department_venues(department_code: str = None, semester: int = None, db: Session = Depends(get_db)):
    query = db.query(models.DepartmentVenueMap)
    if department_code: query = query.filter(models.DepartmentVenueMap.department_code == department_code)
    if semester: query = query.filter(models.DepartmentVenueMap.semester == semester)
    maps = query.all()
    result = []
    for m in maps:
        venue = db.query(models.VenueMaster).filter_by(venue_id=m.venue_id).first()
        if venue:
            result.append(schemas.DepartmentVenueResponse(
                id=m.id, department_code=m.department_code, semester=m.semester,
                venue_id=venue.venue_id, venue_name=venue.venue_name,
                is_lab=venue.is_lab, capacity=venue.capacity
            ))
    return result

@router.post("/department-venues", response_model=schemas.DepartmentVenueResponse)
def create_department_venue(req: schemas.DepartmentVenueCreate, db: Session = Depends(get_db)):
    existing = db.query(models.DepartmentVenueMap).filter_by(
        department_code=req.department_code, semester=req.semester, venue_id=req.venue_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="This venue is already mapped to this department's semester.")
    new_map = models.DepartmentVenueMap(department_code=req.department_code, semester=req.semester, venue_id=req.venue_id)
    db.add(new_map)
    db.commit()
    db.refresh(new_map)
    venue = db.query(models.VenueMaster).filter_by(venue_id=req.venue_id).first()
    return schemas.DepartmentVenueResponse(
        id=new_map.id, department_code=new_map.department_code, semester=new_map.semester,
        venue_id=venue.venue_id, venue_name=venue.venue_name, is_lab=venue.is_lab, capacity=venue.capacity
    )

@router.delete("/department-venues/{map_id}")
def delete_department_venue(map_id: int, db: Session = Depends(get_db)):
    m = db.query(models.DepartmentVenueMap).filter_by(id=map_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Mapping not found")
    db.delete(m)
    db.commit()
    return {"status": "success", "id": map_id}


# ============================================
# COURSE-VENUE MAPPING
# ============================================

@router.get("/course-venues", response_model=List[schemas.CourseVenueResponse])
def get_course_venues(department_code: str = None, db: Session = Depends(get_db)):
    query = db.query(models.CourseVenueMap)
    if department_code: query = query.filter(models.CourseVenueMap.department_code == department_code)
    maps = query.all()
    result = []
    for m in maps:
        venue = db.query(models.VenueMaster).filter_by(venue_id=m.venue_id).first()
        if venue:
            result.append(schemas.CourseVenueResponse(
                id=m.id, department_code=m.department_code, course_code=m.course_code,
                venue_id=venue.venue_id, venue_name=venue.venue_name,
                is_lab=venue.is_lab, capacity=venue.capacity, venue_type=m.venue_type or 'BOTH'
            ))
    return result

@router.post("/course-venues", response_model=schemas.CourseVenueResponse)
def create_course_venue(req: schemas.CourseVenueCreate, db: Session = Depends(get_db)):
    is_common = db.query(models.CommonCourseMap).filter_by(course_code=req.course_code).first()
    if is_common:
        raise HTTPException(status_code=400, detail=f"'{req.course_code}' is a common course. Use Common Course Venues section.")
    existing = db.query(models.CourseVenueMap).filter_by(
        department_code=req.department_code, course_code=req.course_code, venue_id=req.venue_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="This venue is already mapped to the course.")
    new_map = models.CourseVenueMap(
        department_code=req.department_code, course_code=req.course_code,
        venue_id=req.venue_id, venue_type=(req.venue_type or 'BOTH').upper()
    )
    db.add(new_map)
    db.commit()
    db.refresh(new_map)
    venue = db.query(models.VenueMaster).filter_by(venue_id=req.venue_id).first()
    return schemas.CourseVenueResponse(
        id=new_map.id, department_code=new_map.department_code, course_code=new_map.course_code,
        venue_id=venue.venue_id, venue_name=venue.venue_name, is_lab=venue.is_lab,
        capacity=venue.capacity, venue_type=new_map.venue_type or 'BOTH'
    )

@router.delete("/course-venues/{map_id}")
def delete_course_venue(map_id: int, db: Session = Depends(get_db)):
    m = db.query(models.CourseVenueMap).filter_by(id=map_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Mapping not found")
    db.delete(m)
    db.commit()
    return {"status": "success", "id": map_id}


# ============================================
# TIMETABLE GENERATION & RETRIEVAL
# ============================================

DEFAULT_CONFIG = {
    "validation": {"hard_constraint_mode": {"value": False, "enabled": False, "type": "flag", "label": "Hard Constraint Mode", "description": "Block generation when insufficient resources"}},
    "hard_constraints": {
        "max_courses_per_slot": {"value": 1, "enabled": True, "type": "number", "label": "Max Courses Per Slot", "description": "Max courses in a single slot"},
        "lab_block_starts": {"value": [1, 3, 5], "enabled": True, "type": "array", "label": "Lab Block Start Periods", "description": "Periods where lab blocks can begin"},
        "max_lab_blocks_per_day": {"value": 1, "enabled": True, "type": "number", "label": "Max Lab Blocks Per Day", "description": "Max lab blocks per day"},
        "mentor_hour_blocked": {"value": True, "enabled": True, "type": "boolean", "label": "Block Mentor Hour", "description": "Block mentor hour from regular scheduling"},
        "no_faculty_clash": {"value": True, "enabled": True, "type": "boolean", "label": "No Faculty Clash", "description": "No simultaneous faculty scheduling"},
        "no_theory_in_own_lab": {"value": True, "enabled": True, "type": "boolean", "label": "No Theory in Own Lab Block", "description": "No theory during own lab"}
    },
    "dynamic_constraints": {
        "max_theory_per_course_per_day": {"value": 1, "overloaded_value": 2, "enabled": True, "type": "number", "label": "Max Theory/Course/Day", "description": "Max theory per course per day"},
        "max_theory_per_course_per_day_overloaded": {"value": 2, "enabled": True, "type": "number", "label": "Max Theory/Course/Day (Overloaded)", "description": "Max theory when overloaded"},
        "no_back_to_back_theory": {"value": True, "enabled": True, "type": "boolean", "label": "No Back-to-Back Theory", "description": "No consecutive theory"},
        "p8_honours_only": {"value": True, "enabled": True, "type": "boolean", "label": "Period 8 for Honours Only", "description": "Reserve P8 for honours/minor"}
    },
    "soft_constraints": {
        "non_consecutive_lab_days_penalty": {"value": -5, "enabled": True, "type": "number", "label": "Consecutive Lab Days Penalty", "description": "Penalty for consecutive lab days"},
        "theory_lab_same_day_bonus": {"value": 3, "enabled": True, "type": "number", "label": "Theory+Lab Same Day Bonus", "description": "Bonus for same day theory+lab"},
        "fill_slots_bonus": {"value": 10, "enabled": True, "type": "number", "label": "Filled Slot Bonus", "description": "Bonus per filled slot"}
    },
    "section_settings": {
        "min_section_threshold": {"value": 30, "enabled": True, "type": "number", "label": "Min Students for Extra Section", "description": "Min students for extra section"},
        "default_venue_capacity": {"value": 60, "enabled": True, "type": "number", "label": "Default Venue Capacity", "description": "Default venue capacity"}
    },
    "gap_fill": {
        "mini_project_max_periods": {"value": 4, "enabled": True, "type": "number", "label": "Mini Project Max Periods/Week", "description": "Max mini project periods"},
        "core_extra_max_per_week": {"value": 3, "enabled": True, "type": "number", "label": "Core Extra Sessions/Week", "description": "Max extra sessions per week"},
        "core_extra_max_per_day": {"value": 2, "enabled": True, "type": "number", "label": "Core Extra Sessions/Day", "description": "Max extra sessions per day"},
        "open_elective_periods": {"value": 3, "enabled": True, "type": "number", "label": "Open Elective Periods (Sem 5)", "description": "Open elective periods for sem 5"}
    },
    "batch_rotation": {
        "enabled": {"value": True, "enabled": True, "type": "boolean", "label": "Enable Lab Batch Rotation", "description": "Auto rotate lab batches"},
        "venue_aware_rotation": {"value": True, "enabled": True, "type": "boolean", "label": "Venue-Aware Batch Rotation", "description": "Venue-aware rotation"},
        "max_merged_entries": {"value": 15, "enabled": True, "type": "number", "label": "Max Merged Entries Per Slot", "description": "Max entries in merged slot"}
    },
    "elective_handling": {
        "pair_same_category": {"value": True, "enabled": True, "type": "boolean", "label": "Pair Same-Category Electives", "description": "Group same-category electives"},
        "skip_no_faculty_lang": {"value": True, "enabled": True, "type": "boolean", "label": "Skip Language Electives Without Faculty", "description": "Skip unmapped language electives"}
    },
    "honours_minor": {
        "slot_restriction": {"value": 8, "enabled": True, "type": "number", "label": "Honours/Minor Period", "description": "Period for honours/minor courses"}
    }
}

def merge_configs(default_c, saved_c):
    result = copy.deepcopy(default_c)
    if not isinstance(saved_c, dict):
        return result
    for category, items in saved_c.items():
        if category in result and isinstance(items, dict):
            for key, val in items.items():
                if key in result[category] and isinstance(val, dict):
                    for state_key in ['value', 'enabled', 'overloaded_value']:
                        if state_key in val:
                            result[category][key][state_key] = val[state_key]
                else:
                    result[category][key] = val
        else:
            result[category] = items
    return result


@router.post("/generate")
async def generate_timetable(request: schemas.GenerateRequest, db: Session = Depends(get_db)):
    from services.solver_engine import generate_schedule
    import asyncio

    def _get_config():
        return db.query(models.SchedulerConfig).first()

    config_record = await asyncio.to_thread(_get_config)
    config = DEFAULT_CONFIG
    if config_record and config_record.config_json:
        if isinstance(config_record.config_json, str):
            config = json.loads(config_record.config_json)
        else:
            config = config_record.config_json

    hard_mode = False
    try:
        hard_mode_cfg = config.get("validation", {}).get("hard_constraint_mode", {})
        hard_mode = bool(hard_mode_cfg.get("enabled", False))
    except (KeyError, AttributeError):
        pass

    # Process CP-SAT solver in a separate thread so it doesn't block the async event loop
    result = await asyncio.to_thread(
        generate_schedule,
        db,
        request.department_code,
        request.semester,
        request.mentor_day,
        request.mentor_period,
        hard_mode
    )

    if isinstance(result, bool):
        if result:
            return {"status": "success", "message": "Timetable generated successfully", "warnings": []}
        else:
            courses = await run_in_threadpool(lambda: db.query(models.CourseMaster).filter_by(
                department_code=request.department_code, semester=request.semester, is_open_elective=False
            ).all())
            if not courses:
                raise HTTPException(status_code=400, detail=f"No courses found for {request.department_code} Sem {request.semester}.")
            total_ws = sum(c.weekly_sessions for c in courses)
            raise HTTPException(status_code=400, detail=f"Cannot generate: {total_ws} weekly sessions. Schedule may be overloaded.")

    if not result.get("success", False):
        if result.get("errors"):
            raise HTTPException(status_code=422, detail={
                "message": "Resource verification failed.",
                "errors": result.get("errors", []),
                "warnings": result.get("warnings", [])
            })
        else:
            raise HTTPException(status_code=400, detail="Solver failed to find a valid arrangement.")

    return {
        "status": "success",
        "message": f"Successfully generated and saved {result.get('entries_saved', 0)} entries.",
        "warnings": result.get("warnings", []),
        "entries_saved": result.get("entries_saved", 0)
    }


@router.get("/timetable", response_model=List[schemas.TimetableEntry])
def get_timetable(department_code: Optional[str] = None, semester: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(models.TimetableEntry)
    if department_code and department_code.strip():
        query = query.filter(models.TimetableEntry.department_code == department_code)
    if semester:
        query = query.filter(models.TimetableEntry.semester == semester)
    return query.all()

@router.post("/timetable/save")
def save_timetable(request: schemas.TimetableSaveRequest, db: Session = Depends(get_db)):
    db.query(models.TimetableEntry).filter_by(
        department_code=request.department_code, semester=request.semester
    ).delete()
    new_entries = [models.TimetableEntry(**entry.dict()) for entry in request.entries]
    db.add_all(new_entries)
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Save failed: {str(e)}")
    return {"status": "success"}


# ============================================
# CONFLICTS
# ============================================

# Note: Frontend calls ${API_URL}/api/conflicts which becomes /api/api/conflicts
# So we register at /api/conflicts to match 
@router.get("/api/conflicts")
def get_conflicts(department_code: Optional[str] = None, semester: Optional[int] = None, db: Session = Depends(get_db)):
    from services.conflict_detector import detect_conflicts
    return detect_conflicts(db, department_code, semester)

@router.post("/check-conflicts")
def check_conflicts(request: schemas.ConflictCheckRequest, db: Session = Depends(get_db)):
    faculty_conflicts = []
    venue_conflicts = []
    for entry in request.entries:
        day = entry.day_of_week
        period = entry.period_number
        if entry.faculty_name and entry.faculty_name.strip() and entry.faculty_name not in ('', 'Unassigned'):
            busy = db.query(models.TimetableEntry).filter(
                models.TimetableEntry.day_of_week == day,
                models.TimetableEntry.period_number == period,
                models.TimetableEntry.faculty_name == entry.faculty_name,
                ~((models.TimetableEntry.department_code == request.department_code) & (models.TimetableEntry.semester == request.semester))
            ).all()
            for b in busy:
                msg = f"{entry.faculty_name} is occupied by {b.department_code} Sem {b.semester} ({b.course_code}) at {day} Period {period}"
                if msg not in [c['message'] for c in faculty_conflicts]:
                    faculty_conflicts.append({"type": "faculty", "name": entry.faculty_name, "day": day, "period": period, "occupied_by_dept": b.department_code, "occupied_by_sem": b.semester, "occupied_by_course": b.course_code, "message": msg})
        if entry.venue_name and entry.venue_name.strip():
            busy = db.query(models.TimetableEntry).filter(
                models.TimetableEntry.day_of_week == day,
                models.TimetableEntry.period_number == period,
                models.TimetableEntry.venue_name == entry.venue_name,
                ~((models.TimetableEntry.department_code == request.department_code) & (models.TimetableEntry.semester == request.semester))
            ).all()
            for b in busy:
                msg = f"{entry.venue_name} is occupied by {b.department_code} Sem {b.semester} ({b.course_code}) at {day} Period {period}"
                if msg not in [c['message'] for c in venue_conflicts]:
                    venue_conflicts.append({"type": "venue", "name": entry.venue_name, "day": day, "period": period, "occupied_by_dept": b.department_code, "occupied_by_sem": b.semester, "occupied_by_course": b.course_code, "message": msg})
    return {"has_conflicts": len(faculty_conflicts) > 0 or len(venue_conflicts) > 0, "faculty_conflicts": faculty_conflicts, "venue_conflicts": venue_conflicts}


# ============================================
# AVAILABILITY QUERIES (Smart Editor)
# ============================================

@router.get("/available-faculty")
def get_available_faculty(department_code: str = Query(...), day: str = Query(...), period: int = Query(...), course_code: str = None, show_all: bool = False, db: Session = Depends(get_db)):
    busy_names = set()
    busy_entries = db.query(models.TimetableEntry.faculty_name).filter(
        models.TimetableEntry.day_of_week == day, models.TimetableEntry.period_number == period,
        models.TimetableEntry.faculty_name.isnot(None), models.TimetableEntry.faculty_name != '',
        models.TimetableEntry.faculty_name != 'Unassigned'
    ).all()
    for (name,) in busy_entries:
        busy_names.add(name)
    base_faculty = []
    if show_all:
        base_faculty = db.query(models.FacultyMaster).all()
    elif course_code:
        mapped = db.query(models.CourseFacultyMap).filter_by(department_code=department_code, course_code=course_code).all()
        if mapped:
            mapped_ids = [m.faculty_id for m in mapped]
            base_faculty = db.query(models.FacultyMaster).filter(models.FacultyMaster.faculty_id.in_(mapped_ids)).all()
    if not show_all and not base_faculty:
        base_faculty = db.query(models.FacultyMaster).filter_by(department_code=department_code).all()
    result = [{"faculty_id": f.faculty_id, "faculty_name": f.faculty_name, "department_code": f.department_code, "is_available": f.faculty_name not in busy_names} for f in base_faculty]
    result.sort(key=lambda x: (not x["is_available"], x["faculty_name"]))
    return result

@router.get("/available-venues")
def get_available_venues(department_code: str = Query(...), semester: int = Query(...), day: str = Query(...), period: int = Query(...), course_code: str = None, show_all: bool = False, db: Session = Depends(get_db)):
    busy_venue_names = set()
    busy_entries = db.query(models.TimetableEntry.venue_name).filter(
        models.TimetableEntry.day_of_week == day, models.TimetableEntry.period_number == period,
        models.TimetableEntry.venue_name.isnot(None), models.TimetableEntry.venue_name != ''
    ).all()
    for (name,) in busy_entries:
        busy_venue_names.add(name)
    base_venues = []
    if show_all:
        base_venues = db.query(models.VenueMaster).all()
    elif course_code:
        common = db.query(models.CommonCourseMap).filter_by(course_code=course_code, semester=semester).filter(models.CommonCourseMap.venue_name != None).first()
        if common and common.venue_name:
            v_obj = db.query(models.VenueMaster).filter_by(venue_name=common.venue_name).first()
            if v_obj: base_venues = [v_obj]
        if not base_venues:
            mapped = db.query(models.CourseVenueMap).filter_by(department_code=department_code, course_code=course_code).all()
            if mapped:
                mapped_ids = [m.venue_id for m in mapped]
                base_venues = db.query(models.VenueMaster).filter(models.VenueMaster.venue_id.in_(mapped_ids)).all()
    if not show_all and not base_venues:
        dept_venues = db.query(models.DepartmentVenueMap).filter_by(department_code=department_code, semester=semester).all()
        if dept_venues:
            mapped_vids = [v.venue_id for v in dept_venues]
            base_venues = db.query(models.VenueMaster).filter(models.VenueMaster.venue_id.in_(mapped_vids)).all()
        else:
            base_venues = db.query(models.VenueMaster).all()
    result = [{"venue_id": v.venue_id, "venue_name": v.venue_name, "block": v.block, "is_lab": v.is_lab, "capacity": v.capacity, "is_available": v.venue_name not in busy_venue_names} for v in base_venues]
    result.sort(key=lambda x: (not x["is_available"], x["venue_name"]))
    return result


# ============================================
# PERSONALIZED TIMETABLES
# ============================================

@router.get("/timetable/faculty/{faculty_id}")
def get_faculty_timetable(faculty_id: str, db: Session = Depends(get_db)):
    entries = db.query(models.TimetableEntry).filter_by(faculty_id=faculty_id).all()
    slot_map = {}
    conflicts = set()
    for e in entries:
        key = (e.day_of_week, e.period_number)
        slot_map.setdefault(key, []).append(e)
    for key, items in slot_map.items():
        if len(items) > 1:
            venues = set(i.venue_name for i in items if i.venue_name)
            if len(venues) > 1:
                courses_str = " & ".join([f"{i.course_code} ({i.venue_name})" for i in items])
                conflicts.add(f"Conflict on {key[0]} Period {key[1]}: {courses_str}")
    return {"conflicts": list(conflicts), "timetable": entries}

@router.get("/timetable/student/{student_id}")
def get_student_timetable(student_id: str, db: Session = Depends(get_db)):
    student = db.query(models.StudentMaster).filter_by(student_id=student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    regs = db.query(models.CourseRegistration).filter_by(student_id=student_id).all()
    valid_entries = {}
    student_semester = None
    dept_master_slots = []
    for r in regs:
        if student_semester is None:
            student_semester = r.semester
            dept_master_slots = db.query(models.TimetableEntry).filter_by(department_code=student.department_code, semester=student_semester).all()
        c_entries = db.query(models.TimetableEntry).filter_by(course_code=r.course_code, semester=r.semester).all()
        for e in c_entries:
            if e.section_number == 1 or e.section_number is None:
                k = (e.course_code, e.day_of_week, e.period_number)
                if k not in valid_entries:
                    valid_entries[k] = e
    entries = list(valid_entries.values())
    slot_map = {}
    conflicts = set()
    for e in entries:
        key = (e.day_of_week, e.period_number)
        slot_map.setdefault(key, []).append(e)
    for key, items in slot_map.items():
        if len(items) > 1:
            course_codes = set(i.course_code for i in items)
            if len(course_codes) > 1:
                courses_str = " & ".join(course_codes)
                conflicts.add(f"Conflict on {key[0]} Period {key[1]}: {courses_str}")
    return {"conflicts": list(conflicts), "timetable": entries}

@router.get("/timetable/venue/{venue_name}")
def get_venue_timetable(venue_name: str, db: Session = Depends(get_db)):
    all_entries = db.query(models.TimetableEntry).all()
    raw_entries = []
    for e in all_entries:
        if e.venue_name and venue_name in [v.strip() for v in e.venue_name.split(",")]:
            raw_entries.append(e)
    conflicts = set()
    slot_groups = {}
    for e in raw_entries:
        key = (e.day_of_week, e.period_number)
        slot_groups.setdefault(key, []).append(e)
    entries = []
    for slot_key, slot_entries in slot_groups.items():
        course_groups = {}
        for e in slot_entries:
            course_groups.setdefault(e.course_code, []).append(e)
        if len(course_groups) > 1:
            courses_str = " & ".join([f"{code} ({', '.join(set(x.department_code for x in grp))})" for code, grp in course_groups.items()])
            conflicts.add(f"Conflict on {slot_key[0]} Period {slot_key[1]}: {courses_str}")
        for course_code, group in course_groups.items():
            for e in group:
                entries.append(e)
    return {"conflicts": list(conflicts), "timetable": entries}


# ============================================
# STUDENTS & REGISTRATIONS
# ============================================

@router.get("/students", response_model=List[schemas.StudentResponse])
def get_students(department_code: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(models.StudentMaster)
    if department_code:
        query = query.filter_by(department_code=department_code)
    return query.all()

@router.post("/students", response_model=schemas.StudentResponse)
def create_student(req: schemas.StudentCreate, db: Session = Depends(get_db)):
    student = db.query(models.StudentMaster).filter_by(student_id=req.student_id).first()
    if student:
        raise HTTPException(status_code=400, detail="Student ID already exists")
    student = models.StudentMaster(**req.dict())
    db.add(student)
    db.commit()
    db.refresh(student)
    return student

@router.delete("/students/{student_id}")
def delete_student(student_id: str, db: Session = Depends(get_db)):
    student = db.query(models.StudentMaster).filter_by(student_id=student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    regs = db.query(models.CourseRegistration).filter_by(student_id=student_id).all()
    for r in regs:
        db.delete(r)
    db.delete(student)
    db.commit()
    return {"status": "success", "student_id": student_id}

@router.get("/registrations", response_model=List[schemas.CourseRegistrationResponse])
def get_registrations(course_code: Optional[str] = None, semester: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(models.CourseRegistration)
    if course_code: query = query.filter_by(course_code=course_code)
    if semester: query = query.filter_by(semester=semester)
    return query.all()

@router.post("/registrations", response_model=schemas.CourseRegistrationResponse)
def create_registration(req: schemas.CourseRegistrationCreate, db: Session = Depends(get_db)):
    student = db.query(models.StudentMaster).filter_by(student_id=req.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    existing = db.query(models.CourseRegistration).filter_by(student_id=req.student_id, course_code=req.course_code, semester=req.semester).first()
    if existing:
        raise HTTPException(status_code=400, detail="Student already registered for this course")
    reg = models.CourseRegistration(**req.dict())
    db.add(reg)
    db.commit()
    db.refresh(reg)
    return reg

@router.delete("/registrations/{reg_id}")
def delete_registration(reg_id: int, db: Session = Depends(get_db)):
    reg = db.query(models.CourseRegistration).filter_by(id=reg_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
    db.delete(reg)
    db.commit()
    return {"status": "success", "id": reg_id}


# ============================================
# SCHEDULER CONFIG
# ============================================
# Note: Frontend calls ${API_URL}/api/config → /api/api/config
# So we register at /api/config

@router.get("/api/config")
def get_scheduler_config(db: Session = Depends(get_db)):
    row = db.query(models.SchedulerConfig).filter_by(id=1).first()
    if not row:
        row = models.SchedulerConfig(id=1, config_json=json.dumps(DEFAULT_CONFIG))
        db.add(row)
        db.commit()
        db.refresh(row)
        return json.loads(row.config_json)
    saved_config = json.loads(row.config_json) if row.config_json else {}
    merged_config = merge_configs(DEFAULT_CONFIG, saved_config)
    return merged_config

@router.put("/api/config")
def update_scheduler_config(config: dict, db: Session = Depends(get_db)):
    row = db.query(models.SchedulerConfig).filter_by(id=1).first()
    if not row:
        row = models.SchedulerConfig(id=1, config_json=json.dumps(config))
        db.add(row)
    else:
        row.config_json = json.dumps(config)
    db.commit()
    return {"status": "saved"}

@router.post("/api/config/reset")
def reset_scheduler_config(db: Session = Depends(get_db)):
    row = db.query(models.SchedulerConfig).filter_by(id=1).first()
    if row:
        row.config_json = json.dumps(DEFAULT_CONFIG)
    else:
        row = models.SchedulerConfig(id=1, config_json=json.dumps(DEFAULT_CONFIG))
        db.add(row)
    db.commit()
    return json.loads(row.config_json)


# ============================================
# SEMESTER CONFIG (with /api/ prefix for frontend)
# ============================================
# Frontend calls ${API_URL}/api/semester-config → /api/api/semester-config

@router.get("/api/semester-config", response_model=List[schemas.SemesterConfigResponse])
def get_semester_configs_api(db: Session = Depends(get_db)):
    return db.query(models.SemesterConfig).all()

@router.post("/api/semester-config/{semester}", response_model=schemas.SemesterConfigResponse)
def update_semester_config_api(semester: int, req: schemas.SemesterConfigUpdate, db: Session = Depends(get_db)):
    config = db.query(models.SemesterConfig).filter_by(semester=semester).first()
    if not config:
        config = models.SemesterConfig(semester=semester, academic_year=req.academic_year)
        db.add(config)
    else:
        config.academic_year = req.academic_year
    db.commit()
    db.refresh(config)
    return config


# ============================================
# COMMON COURSES
# ============================================

@router.get("/common-courses")
def get_common_courses(db: Session = Depends(get_db)):
    rows = db.query(models.CommonCourseMap).all()
    groups = {}
    for row in rows:
        key = (row.course_code, row.semester)
        if key not in groups:
            groups[key] = {"course_code": row.course_code, "semester": row.semester, "departments": [], "venue_name": row.venue_name, "venue_type": row.venue_type or 'BOTH', "dept_student_counts": []}
        sem_count = db.query(models.DepartmentSemesterCount).filter_by(department_code=row.department_code, semester=row.semester).first()
        count = sem_count.student_count if sem_count and sem_count.student_count > 0 else 0
        groups[key]["departments"].append(row.department_code)
        groups[key]["dept_student_counts"].append({"dept": row.department_code, "count": count})
        if row.venue_name and not groups[key]["venue_name"]:
            groups[key]["venue_name"] = row.venue_name
            groups[key]["venue_type"] = row.venue_type or 'BOTH'
    return list(groups.values())


class CommonCourseRequest(BaseModel):
    course_code: str
    semester: int
    department_codes: List[str]

@router.post("/common-courses")
def save_common_course(req: CommonCourseRequest, db: Session = Depends(get_db)):
    existing_venue = None
    existing_type = 'BOTH'
    old_rows = db.query(models.CommonCourseMap).filter_by(course_code=req.course_code, semester=req.semester).all()
    for r in old_rows:
        if r.venue_name:
            existing_venue = r.venue_name
            existing_type = r.venue_type or 'BOTH'
            break
    db.query(models.CommonCourseMap).filter_by(course_code=req.course_code, semester=req.semester).delete()
    for dept in req.department_codes:
        db.add(models.CommonCourseMap(course_code=req.course_code, semester=req.semester, department_code=dept, venue_name=existing_venue, venue_type=existing_type))
    db.commit()
    return {"status": "saved", "course_code": req.course_code, "semester": req.semester}


class CommonCourseVenueRequest(BaseModel):
    course_code: str
    semester: int
    venue_name: str
    venue_type: str = 'BOTH'

@router.post("/common-courses/venue")
def set_common_course_venue(req: CommonCourseVenueRequest, db: Session = Depends(get_db)):
    rows = db.query(models.CommonCourseMap).filter_by(course_code=req.course_code, semester=req.semester).all()
    if not rows:
        raise HTTPException(status_code=404, detail=f"No common course found for {req.course_code} semester {req.semester}")
    vtype = (req.venue_type or 'BOTH').upper()
    db.query(models.CommonCourseMap).filter_by(course_code=req.course_code, semester=req.semester).update({"venue_name": req.venue_name, "venue_type": vtype})
    db.query(models.TimetableEntry).filter_by(course_code=req.course_code, semester=req.semester).update({"venue_name": req.venue_name})
    db.commit()
    dept_codes = [r.department_code for r in rows]
    return {"status": "saved", "course_code": req.course_code, "semester": req.semester, "venue_name": req.venue_name, "venue_type": vtype, "synced_departments": dept_codes}

@router.delete("/common-courses/venue/{course_code:path}/{semester}")
def clear_common_course_venue(course_code: str, semester: int, db: Session = Depends(get_db)):
    affected = db.query(models.CommonCourseMap).filter_by(course_code=course_code, semester=semester).update({"venue_name": None, "venue_type": "BOTH"})
    db.commit()
    return {"status": "cleared", "rows": affected}

@router.get("/common-courses/student-distribution/{course_code:path}/{semester}")
def get_common_course_students(course_code: str, semester: int, db: Session = Depends(get_db)):
    rows = db.query(models.CommonCourseMap).filter_by(course_code=course_code, semester=semester).all()
    if not rows:
        raise HTTPException(status_code=404, detail="Common course not found")
    breakdown = []
    total = 0
    for row in rows:
        sem_count = db.query(models.DepartmentSemesterCount).filter_by(department_code=row.department_code, semester=semester).first()
        count = sem_count.student_count if sem_count and sem_count.student_count > 0 else 0
        breakdown.append({"dept": row.department_code, "count": count})
        total += count
    return {"course_code": course_code, "semester": semester, "total": total, "departments": breakdown}

@router.delete("/common-courses/{course_code:path}/{semester}")
def delete_common_course(course_code: str, semester: int, db: Session = Depends(get_db)):
    deleted = db.query(models.CommonCourseMap).filter_by(course_code=course_code, semester=semester).delete()
    db.commit()
    return {"status": "deleted", "rows": deleted}


# ============================================
# USER CONSTRAINTS
# ============================================
# Frontend calls ${API_URL}/api/user-constraints → /api/api/user-constraints

def _constraint_to_response(row):
    return {
        "id": row.id, "uuid": row.uuid, "name": row.name, "description": row.description,
        "enabled": row.enabled, "priority": row.priority, "soft_weight": row.soft_weight,
        "constraint_type": row.constraint_type, "scope": json.loads(row.scope_json),
        "target": json.loads(row.target_json), "rules": json.loads(row.rules_json),
        "created_at": row.created_at, "updated_at": row.updated_at, "order_index": row.order_index,
    }

@router.get("/api/user-constraints")
def list_user_constraints(dept: Optional[str] = None, sem: Optional[int] = None, db: Session = Depends(get_db)):
    rows = db.query(models.UserConstraint).order_by(models.UserConstraint.priority.desc(), models.UserConstraint.order_index).all()
    results = []
    for row in rows:
        resp = _constraint_to_response(row)
        if dept or sem:
            scope = resp["scope"]
            scope_depts = scope.get("departments", ["*"])
            scope_sems = scope.get("semesters", ["*"])
            if dept and scope_depts != ["*"] and dept not in scope_depts: continue
            if sem is not None and scope_sems != ["*"] and sem not in scope_sems: continue
        results.append(resp)
    return results

@router.post("/api/user-constraints")
def create_user_constraint(data: schemas.UserConstraintCreate, db: Session = Depends(get_db)):
    now = datetime.utcnow().isoformat()
    max_order = db.query(models.UserConstraint).count()
    row = models.UserConstraint(
        uuid=str(uuid_lib.uuid4()), name=data.name, description=data.description,
        enabled=data.enabled, priority=data.priority, soft_weight=data.soft_weight,
        constraint_type=data.constraint_type, scope_json=json.dumps(data.scope),
        target_json=json.dumps(data.target), rules_json=json.dumps(data.rules),
        created_at=now, updated_at=now, order_index=max_order,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _constraint_to_response(row)

@router.put("/api/user-constraints/{constraint_uuid}")
def update_user_constraint(constraint_uuid: str, data: schemas.UserConstraintUpdate, db: Session = Depends(get_db)):
    row = db.query(models.UserConstraint).filter_by(uuid=constraint_uuid).first()
    if not row:
        raise HTTPException(status_code=404, detail="Constraint not found")
    now = datetime.utcnow().isoformat()
    if data.name is not None: row.name = data.name
    if data.description is not None: row.description = data.description
    if data.enabled is not None: row.enabled = data.enabled
    if data.priority is not None: row.priority = data.priority
    if data.soft_weight is not None: row.soft_weight = data.soft_weight
    if data.constraint_type is not None: row.constraint_type = data.constraint_type
    if data.scope is not None: row.scope_json = json.dumps(data.scope)
    if data.target is not None: row.target_json = json.dumps(data.target)
    if data.rules is not None: row.rules_json = json.dumps(data.rules)
    row.updated_at = now
    db.commit()
    db.refresh(row)
    return _constraint_to_response(row)

@router.delete("/api/user-constraints/{constraint_uuid}")
def delete_user_constraint(constraint_uuid: str, db: Session = Depends(get_db)):
    deleted = db.query(models.UserConstraint).filter_by(uuid=constraint_uuid).delete()
    if not deleted:
        raise HTTPException(status_code=404, detail="Constraint not found")
    db.commit()
    return {"status": "deleted", "uuid": constraint_uuid}

@router.patch("/api/user-constraints/{constraint_uuid}/toggle")
def toggle_user_constraint(constraint_uuid: str, db: Session = Depends(get_db)):
    row = db.query(models.UserConstraint).filter_by(uuid=constraint_uuid).first()
    if not row:
        raise HTTPException(status_code=404, detail="Constraint not found")
    row.enabled = not row.enabled
    row.updated_at = datetime.utcnow().isoformat()
    db.commit()
    db.refresh(row)
    return _constraint_to_response(row)

@router.post("/api/user-constraints/reorder")
def reorder_user_constraints(data: schemas.UserConstraintReorderRequest, db: Session = Depends(get_db)):
    for idx, c_uuid in enumerate(data.order):
        row = db.query(models.UserConstraint).filter_by(uuid=c_uuid).first()
        if row:
            row.order_index = idx
    db.commit()
    return {"status": "reordered"}

@router.post("/api/user-constraints/validate")
def validate_user_constraint(data: schemas.UserConstraintCreate, db: Session = Depends(get_db)):
    warnings = []
    valid_types = ['COURSE_INJECTION', 'SLOT_BLOCKING', 'FACULTY_RULE', 'SPACING_RULE', 'DISTRIBUTION_RULE', 'CAPACITY_RULE', 'CUSTOM_PLACEMENT']
    if data.constraint_type not in valid_types:
        warnings.append(f"Unknown constraint type: {data.constraint_type}")
    valid_priorities = ['HARD', 'SOFT', 'PREFERENCE']
    if data.priority not in valid_priorities:
        warnings.append(f"Unknown priority: {data.priority}")
    scope = data.scope or {}
    scope_depts = scope.get("departments", ["*"])
    if scope_depts != ["*"]:
        for dept in scope_depts:
            if not db.query(models.DepartmentMaster).filter_by(department_code=dept).first():
                warnings.append(f"Department '{dept}' not found")
    return {"valid": len(warnings) == 0, "warnings": warnings}


# ============================================
# EXPORT
# ============================================

@router.get("/export/timetable/excel")
def export_timetable_excel(department_code: str = None, semester: int = None, db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    slots = db.query(models.SlotMaster).order_by(models.SlotMaster.day_of_week, models.SlotMaster.period_number).all()
    if not slots:
        raise HTTPException(status_code=400, detail="No slots configured")

    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    active_days = [d for d in days_order if any(s.day_of_week == d for s in slots)]
    if not active_days: active_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    max_period = max([s.period_number for s in slots], default=8)
    periods = list(range(1, max_period + 1))

    query = db.query(models.TimetableEntry)
    if department_code: query = query.filter(models.TimetableEntry.department_code == department_code)
    if semester: query = query.filter(models.TimetableEntry.semester == semester)
    entries = query.all()

    timetable_dict = {}
    for e in entries:
        key = (e.day_of_week, e.period_number)
        timetable_dict.setdefault(key, []).append(e)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"

    thin = Side(border_style="thin", color="D1D5DB")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
    fill_header = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_day = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_lab = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    fill_class = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_empty = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    font_bold = Font(bold=True, color="1F2937", size=11)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    header_row = 2
    d_cell = ws.cell(row=header_row, column=1, value="DAY / PERIOD")
    d_cell.font = font_bold
    d_cell.alignment = align_center
    d_cell.fill = fill_day
    d_cell.border = border_all
    ws.column_dimensions['A'].width = 16

    for i, p in enumerate(periods):
        col = i + 2
        slot = next((s for s in slots if s.period_number == p and s.day_of_week == 'Monday'), None)
        time_str = f"{slot.start_time} - {slot.end_time}" if slot else ""
        header_text = f"Period {p}\n({time_str})"
        c = ws.cell(row=header_row, column=col, value=header_text)
        c.font = Font(bold=True, color="374151", size=10)
        c.alignment = align_center
        c.fill = fill_header
        c.border = border_all
        ws.column_dimensions[get_column_letter(col)].width = 24

    current_row = header_row + 1
    for day in active_days:
        dc = ws.cell(row=current_row, column=1, value=day[:3].upper())
        dc.font = Font(bold=True, size=12, color="1F2937")
        dc.alignment = align_center
        dc.fill = fill_day
        dc.border = border_all
        ws.row_dimensions[current_row].height = 85

        for i, p in enumerate(periods):
            col = i + 2
            cell_entries = timetable_dict.get((day, p), [])
            if not cell_entries:
                cell_text = "Empty Slot"
                fill = fill_empty
                cell_font = Font(italic=True, color="9CA3AF", size=10)
            else:
                is_lab = any(e.session_type and e.session_type.upper() == 'LAB' for e in cell_entries)
                lines = []
                for e in cell_entries:
                    block = f"{e.course_code}\n{e.course_name or ''}\n({e.faculty_name or 'Unassigned'})\n[{e.venue_name or ''}]"
                    lines.append(block)
                cell_text = "\n---\n".join(lines)
                fill = fill_lab if is_lab else fill_class
                cell_font = Font(bold=False, size=10, color="111827")

            c = ws.cell(row=current_row, column=col, value=cell_text)
            c.alignment = align_center
            c.fill = fill
            c.border = border_all
            c.font = cell_font
        current_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    headers = {'Content-Disposition': f'attachment; filename="timetable_{department_code or "all"}_{semester or "all"}.xlsx"'}
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
