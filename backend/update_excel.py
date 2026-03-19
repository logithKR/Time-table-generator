import sys

file_path = r'c:\Users\kalai\Downloads\time table\backend\main.py'

with open(file_path, 'r', encoding='utf-8') as f:
    content = f.read()

target = '''    # Title Row
    title = f"TIMETABLE - {department_code or 'ALL DEPARTMENTS'} (SEMESTER {semester or 'ALL'})"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(periods)+1)
    cell = ws.cell(row=1, column=1, value=title.upper())
    cell.font = font_title
    cell.alignment = align_center
    ws.row_dimensions[1].height = 40'''

# let's try to find it by stripping whitespace to be safe
content_stripped = content.replace('\r\n', '\n')
target_stripped = target.replace('\r\n', '\n')

replacement = '''    try:
        from openpyxl.drawing.image import Image
    except ImportError:
        Image = None
    import os

    # Fetch Academic Year
    academic_year = "___________"
    if semester:
        config = db.query(models.SemesterConfig).filter_by(semester=semester).first()
        if config and config.academic_year:
            academic_year = config.academic_year
    
    # Fetch active breaks
    all_breaks = db.query(models.BreakMaster).all()
    active_breaks = [b for b in all_breaks if not b.semester_ids or (semester and semester in b.semester_ids)]

    fill_break = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    header_cols_total = len(periods) + 1

    # --- ROW 1-5: INSTITUTION HEADER ---
    logo_path = os.path.join(os.path.dirname(__file__), "bitsathy-logo.png")
    if Image and os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            img.width = 100
            img.height = 100
            ws.add_image(img, "A1")
        except Exception as e:
            print(f"Warning: Could not inject logo: {e}")

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=header_cols_total)
    c1 = ws.cell(row=1, column=2, value="BANNARI AMMAN INSTITUTE OF TECHNOLOGY")
    c1.font = Font(bold=True, size=18, color="1E3A8A")
    c1.alignment = align_center

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=header_cols_total)
    c2 = ws.cell(row=2, column=2, value="(An Autonomous Institution Affiliated to Anna University, Chennai)")
    c2.font = Font(bold=True, size=11, color="111827")
    c2.alignment = align_center

    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=header_cols_total)
    c3 = ws.cell(row=3, column=2, value="Sathyamangalam - 638401")
    c3.font = Font(bold=True, size=11, color="111827")
    c3.alignment = align_center

    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=header_cols_total)
    c4 = ws.cell(row=4, column=2, value="CLASS TIMETABLE")
    c4.font = Font(bold=True, size=14, color="111827")
    c4.alignment = align_center

    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=header_cols_total)
    c5 = ws.cell(row=5, column=2, value=f"Academic Year {academic_year}")
    c5.font = Font(bold=True, size=12, color="111827")
    c5.alignment = align_center

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20

    # --- ROW 6-7: BREAK DISPLAY HEADER ---
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=header_cols_total)
    b_head = ws.cell(row=6, column=1, value="SCHEDULED BREAKS")
    b_head.font = Font(bold=True, size=10, color="1F2937")
    b_head.fill = fill_day
    b_head.alignment = align_center
    b_head.border = border_all

    break_text = "  |  ".join([f"{b.break_type}: {b.start_time}-{b.end_time}" for b in active_breaks]) if active_breaks else "No breaks scheduled"
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=header_cols_total)
    b_val = ws.cell(row=7, column=1, value=break_text)
    b_val.font = Font(bold=True, size=10, color="4B5563")
    b_val.fill = fill_break
    b_val.alignment = align_center
    b_val.border = border_all

    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 25

    # --- ROW 8: DEPARTMENT AND SEMESTER META ROW ---
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=header_cols_total)
    meta_text = f"DEPARTMENT OF {department_code or 'ALL'} | SEMESTER {semester or 'ALL'}"
    m_cell = ws.cell(row=8, column=1, value=meta_text)
    m_cell.font = Font(bold=True, size=12, color="000000")
    m_cell.alignment = align_center
    ws.row_dimensions[8].height = 25'''

if target_stripped in content_stripped:
    new_content = content_stripped.replace(target_stripped, replacement)
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(new_content)
    print("SUCCESS")
else:
    print("FAILED TO FIND TARGET")
